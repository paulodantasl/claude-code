#!/usr/bin/env python3
"""Build a formatted, formula-driven estimate.xlsx from a project's CSV inputs.

Usage:
    python estimating/scripts/build_estimate_xlsx.py <project_dir>

Reads:
    <project_dir>/lineitems.csv   (schema in estimating/templates/estimate-workbook.md)
    <project_dir>/markups.csv     (key,value percent pairs)

Writes:
    <project_dir>/estimate.xlsx   (Detail + Summary sheets, live Excel formulas)

The workbook stays formula-driven so the auditor (or Excel) can change a quantity or
unit cost and see totals recompute. Material sales tax applies to material only; the
remaining markups cascade once, in order.
"""

import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl  (it is in requirements.txt)")

CURRENCY = '#,##0.00'
PCT = '0.0"%"'

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBTOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Markup waterfall: (csv_key, label). Applied once, in order, to the running subtotal.
# Material sales tax is handled separately (material extensions only).
MARKUP_ORDER = [
    ("general_conditions_pct", "General Conditions"),
    ("contingency_pct", "Contingency / Escalation"),
    ("insurance_pct", "Insurance (GL / Builder's Risk)"),
    ("bond_pct", "Payment & Performance Bond"),
    ("permit_pct", "Permit & Fees"),
    ("ohp_pct", "Overhead & Profit"),
]


def _num(value, default=0.0):
    try:
        s = str(value).strip().replace("$", "").replace(",", "").replace("%", "")
        return float(s) if s else default
    except (TypeError, ValueError):
        return default


def read_lineitems(path: Path):
    rows = []
    with path.open(newline="", encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            # skip wholly empty rows
            if not any((v or "").strip() for v in r.values()):
                continue
            rows.append(r)
    if not rows:
        sys.exit(f"No line items found in {path}")
    # stable sort by division then section so the workbook reads cleanly
    rows.sort(key=lambda r: (str(r.get("division", "")).zfill(3),
                             str(r.get("section", ""))))
    return rows


def read_markups(path: Path):
    mk = {}
    if path.exists():
        with path.open(newline="", encoding="utf-8-sig") as fh:
            for r in csv.reader(fh):
                if len(r) >= 2 and r[0].strip() and not r[0].strip().startswith("#"):
                    mk[r[0].strip()] = _num(r[1])
    return mk


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def build_detail(wb, items):
    ws = wb.active
    ws.title = "Detail"
    headers = ["Div", "Section", "Item", "Description", "Qty", "Unit",
               "Unit Mat", "Waste %", "Mat Ext", "Unit Lab", "Lab Ext",
               "Unit Equip", "Equip Ext", "Unit Sub", "Sub Ext", "Line Total", "Notes"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for it in items:
        r = ws.max_row + 1
        ws.cell(r, 1, str(it.get("division", "")).strip())
        ws.cell(r, 2, str(it.get("section", "")).strip())
        ws.cell(r, 3, it.get("item", ""))
        ws.cell(r, 4, it.get("description", ""))
        ws.cell(r, 5, _num(it.get("qty")))
        ws.cell(r, 6, str(it.get("unit", "")).strip())
        ws.cell(r, 7, _num(it.get("unit_mat")))
        ws.cell(r, 8, _num(it.get("waste_pct")))
        ws.cell(r, 9, f"=E{r}*G{r}*(1+H{r}/100)")     # Mat Ext (pre-tax)
        ws.cell(r, 10, _num(it.get("unit_lab")))
        ws.cell(r, 11, f"=E{r}*J{r}")                  # Lab Ext
        ws.cell(r, 12, _num(it.get("unit_equip")))
        ws.cell(r, 13, f"=E{r}*L{r}")                  # Equip Ext
        ws.cell(r, 14, _num(it.get("unit_sub")))
        ws.cell(r, 15, f"=E{r}*N{r}")                  # Sub Ext
        ws.cell(r, 16, f"=I{r}+K{r}+M{r}+O{r}")        # Line Total
        ws.cell(r, 17, it.get("notes", ""))

    last = ws.max_row
    for r in range(2, last + 1):
        for c in (7, 9, 10, 11, 12, 13, 14, 15, 16):
            ws.cell(r, c).number_format = CURRENCY
        ws.cell(r, 8).number_format = PCT
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).border = BORDER

    widths = [6, 11, 22, 40, 9, 7, 11, 8, 13, 11, 13, 11, 13, 11, 13, 13, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return last


def build_summary(wb, items, last_detail_row, markups):
    ws = wb.create_sheet("Summary", 0)
    d = "Detail!"
    rng = lambda col: f"{d}${col}$2:${col}${last_detail_row}"

    ws["A1"] = "ESTIMATE SUMMARY"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Project:"
    ws["B2"] = "{{PROJECT_NAME}} — {{CITY, COUNTY, FL}}"
    ws["A3"] = "All figures are formula-linked to the Detail sheet."
    ws["A3"].font = Font(italic=True, size=9, color="808080")

    # Cost by CSI division
    hr = 5
    ws.cell(hr, 1, "Cost by CSI Division")
    ws.cell(hr, 1).font = Font(bold=True)
    cols = ["Div", "Material", "Labor", "Equipment", "Subcontract", "Total"]
    for i, h in enumerate(cols, start=1):
        ws.cell(hr + 1, i, h)
    style_header(ws, hr + 1, len(cols))

    divisions = sorted({str(it.get("division", "")).strip() for it in items if str(it.get("division", "")).strip()},
                       key=lambda x: x.zfill(3))
    row = hr + 2
    first_div_row = row
    for div in divisions:
        ws.cell(row, 1, div)
        ws.cell(row, 2, f'=SUMIF({rng("A")},A{row},{rng("I")})')
        ws.cell(row, 3, f'=SUMIF({rng("A")},A{row},{rng("K")})')
        ws.cell(row, 4, f'=SUMIF({rng("A")},A{row},{rng("M")})')
        ws.cell(row, 5, f'=SUMIF({rng("A")},A{row},{rng("O")})')
        ws.cell(row, 6, f"=SUM(B{row}:E{row})")
        for c in range(2, 7):
            ws.cell(row, c).number_format = CURRENCY
            ws.cell(row, c).border = BORDER
        ws.cell(row, 1).border = BORDER
        row += 1
    last_div_row = row - 1

    # Direct cost totals
    ws.cell(row, 1, "Direct Cost")
    ws.cell(row, 1).font = Font(bold=True)
    for c, col in ((2, "B"), (3, "C"), (4, "D"), (5, "E"), (6, "F")):
        ws.cell(row, c, f"=SUM({col}{first_div_row}:{col}{last_div_row})")
        ws.cell(row, c).number_format = CURRENCY
        ws.cell(row, c).font = Font(bold=True)
        ws.cell(row, c).fill = SUBTOTAL_FILL
    direct_row = row
    total_material_ref = f"B{direct_row}"   # total material (pre-tax) for sales tax
    running = f"F{direct_row}"              # running subtotal cell ref

    # Markup waterfall
    row += 2
    ws.cell(row, 1, "Markups / Cost Build-up")
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    ws.cell(row, 1, "Item"); ws.cell(row, 5, "Rate"); ws.cell(row, 6, "Amount")
    style_header(ws, row, 6)

    def add_line(label, formula, rate_cell_pct=None, bold=False, fill=None):
        nonlocal row
        row += 1
        ws.cell(row, 1, label)
        if rate_cell_pct is not None:
            ws.cell(row, 5, rate_cell_pct)
            ws.cell(row, 5).number_format = PCT
        ws.cell(row, 6, formula)
        ws.cell(row, 6).number_format = CURRENCY
        if bold:
            ws.cell(row, 1).font = Font(bold=True)
            ws.cell(row, 6).font = Font(bold=True)
        if fill:
            for c in range(1, 7):
                ws.cell(row, c).fill = fill
        return row

    add_line("Direct Cost", f"={running}", bold=True)
    subtotal_ref = f"F{row}"

    tax_pct = markups.get("material_sales_tax_pct", 0.0)
    r = add_line("Material Sales Tax (materials only)",
                 f"={total_material_ref}*{tax_pct}/100", rate_cell_pct=tax_pct)
    add_line("Subtotal", f"={subtotal_ref}+F{r}", bold=True)
    subtotal_ref = f"F{row}"

    for key, label in MARKUP_ORDER:
        pct = markups.get(key, 0.0)
        r = add_line(label, f"={subtotal_ref}*{pct}/100", rate_cell_pct=pct)
        add_line("Subtotal", f"={subtotal_ref}+F{r}", bold=True)
        subtotal_ref = f"F{row}"

    bid_row = add_line("BID TOTAL", f"={subtotal_ref}", bold=True, fill=TOTAL_FILL)
    ws.cell(bid_row, 1, "BID TOTAL")
    ws.cell(bid_row, 6).font = Font(bold=True, size=12)

    for w, i in ((38, 1), (14, 2), (14, 3), (14, 4), (14, 5), (16, 6)):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def main():
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    proj = Path(sys.argv[1])
    if not proj.is_dir():
        sys.exit(f"Not a directory: {proj}")

    items = read_lineitems(proj / "lineitems.csv")
    markups = read_markups(proj / "markups.csv")

    wb = Workbook()
    last = build_detail(wb, items)
    build_summary(wb, items, last, markups)

    out = proj / "estimate.xlsx"
    wb.save(out)
    print(f"Wrote {out}  ({len(items)} line items)")


if __name__ == "__main__":
    main()
