#!/usr/bin/env python3
"""Deterministic estimate validator — machine checks for the accuracy protocol.

Usage:
    python3 estimating/scripts/validate_estimate.py <project_dir> [--sector SECTOR]
                                                     [--strict] [--xlsx PATH]

    SECTOR: residential | commercial | ti | public   (default: residential)
    --strict: exit 1 on WARNs too (CI gate)
    --xlsx:   path to estimate.xlsx (default <project_dir>/estimate.xlsx)

Checks lineitems.csv + markups.csv (+ scope-of-work.md if present):
  schema, numeric parse, unit whitelist, zero-qty dispositions, rollup pricing
  guard, waste placement, mixed self-perform/sub lines, markup sanity ranges,
  waterfall recompute, division benchmark bands by sector, missing-division
  sweep, duplicate lines, division-code validity, qty outliers, best-effort
  allowance tie-out vs the scope, and — when estimate.xlsx and openpyxl are
  available — a structural tie-out of the workbook itself (constants, extension
  formulas, subtotal ranges, waterfall order/rates vs the CSVs).

Exit code 0 = no FAILs (WARNs allowed; --strict fails WARNs too), 1 = FAIL.
Prompt discipline catches most errors; this catches the rest for free.
"""

import argparse
import csv
import re
import sys
from statistics import median
from collections import defaultdict
from pathlib import Path

HEADER = ["division", "section", "item", "description", "qty", "unit",
          "unit_mat", "unit_lab", "unit_equip", "unit_sub", "waste_pct", "notes"]

# CSI MasterFormat division codes accepted in lineitems.csv (zero-padded 01-49).
CSI_DIVS = {f"{i:02d}" for i in range(1, 50)}

# Must mirror build_estimate_xlsx.MARKUP_ORDER (keys and labels) for the xlsx tie-out.
MARKUP_ORDER = [
    ("general_conditions_pct", "General Conditions"),
    ("contingency_pct", "Contingency / Escalation"),
    ("insurance_pct", "Insurance (GL / Builder's Risk)"),
    ("bond_pct", "Payment & Performance Bond"),
    ("permit_pct", "Permit & Fees"),
    ("ohp_pct", "Overhead & Profit"),
]

UNIT_WHITELIST = {
    "EA", "LF", "SF", "SY", "CY", "SQ", "TON", "TONS", "LBS", "LB", "LS",
    "ALLOW", "MBF", "BF", "GAL", "HR", "DAY", "WK", "MO", "SET", "PR", "CF",
    "RL", "BX", "CT", "KIT",
}

# Division benchmark bands: {sector: {div: (lo_pct, hi_pct)}} — % of direct cost.
# estimating-accuracy-protocol.md §1 bands widened by a WARN tolerance (a screening
# net, intentionally looser than the protocol table — not a mirror). WARN outside.
BANDS = {
    "residential": {
        "01": (0, 12), "03": (5, 20), "04": (1, 10), "05": (0.5, 8), "06": (4, 16),
        "07": (3, 10), "08": (3, 13), "09": (10, 26), "21": (0, 2), "22": (3, 9),
        "23": (4, 10), "26": (4, 9), "27": (0, 3), "28": (0, 2), "31": (2, 16),
        "32": (1, 8), "33": (0, 4),
    },
    "commercial": {
        "01": (2, 14), "03": (6, 26), "04": (1, 12), "05": (3, 14), "06": (0.5, 8),
        "07": (3, 10), "08": (3, 12), "09": (6, 17), "21": (1, 4), "22": (3, 9),
        "23": (6, 14), "26": (7, 15), "27": (0.5, 4), "28": (0.5, 4), "31": (4, 16),
        "32": (2, 10), "33": (1, 8),
    },
    "ti": {
        "01": (4, 17), "02": (1, 10), "03": (0, 6), "04": (0, 4), "05": (0.5, 6),
        "06": (1, 10), "07": (0.5, 6), "08": (3, 12), "09": (16, 38), "21": (1, 6),
        "22": (2, 12), "23": (8, 20), "26": (8, 20), "27": (1, 6), "28": (1, 6),
        "31": (0, 3), "32": (0, 3),
    },
    # public = civil/vertical mix varies too much for tight bands; use wide commercial bands
    "public": {
        "01": (2, 16), "03": (4, 30), "05": (2, 16), "09": (4, 18), "22": (2, 10),
        "23": (4, 15), "26": (6, 16), "31": (4, 30), "32": (2, 14), "33": (1, 14),
    },
}

# Divisions that should exist (priced or explicitly excluded) per sector → WARN if absent.
EXPECTED_DIVS = {
    "residential": {"03", "06", "07", "08", "09", "22", "23", "26", "31"},
    "commercial": {"03", "05", "07", "08", "09", "21", "22", "23", "26", "28", "31", "32"},
    "ti": {"01", "08", "09", "22", "23", "26", "28"},
    "public": {"01", "03", "26", "31"},
}

MARKUP_SANE = {  # key: (lo, hi) in percent — WARN outside
    "material_sales_tax_pct": (0, 9), "general_conditions_pct": (0, 25),
    "contingency_pct": (0, 20), "insurance_pct": (0, 5), "bond_pct": (0, 4),
    "permit_pct": (0, 6), "ohp_pct": (0, 25),
}

ROLLUP_RE = re.compile(r"\b(total|rollup|roll-up|subtotal)\b", re.I)
INFO_NOTE_RE = re.compile(r"\b(rollup|roll-up|info|sanity|do not double|dedup)\b", re.I)
ZERO_OK_NOTE_RE = re.compile(r"\b(rfi|deferred|by others|allowance|see |qty tbd|verify)\b", re.I)


def num(s, default=0.0):
    try:
        s = str(s).strip().replace("$", "").replace(",", "").replace("%", "")
        return float(s) if s else default
    except (TypeError, ValueError):
        return None  # signal parse failure distinctly


class Report:
    def __init__(self):
        self.rows = []

    def add(self, level, check, msg):
        self.rows.append((level, check, msg))

    def dump(self):
        order = {"FAIL": 0, "WARN": 1, "INFO": 2, "PASS": 3}
        counts = defaultdict(int)
        for level, check, msg in sorted(self.rows, key=lambda r: order[r[0]]):
            counts[level] += 1
            print(f"  [{level}] {check}: {msg}")
        print(f"\n  Summary: {counts['FAIL']} FAIL / {counts['WARN']} WARN / "
              f"{counts['INFO']} INFO / {counts['PASS']} PASS")
        return counts


def tie_out_xlsx(xlsx_path, data, mk, rep):
    """Structurally verify estimate.xlsx against the CSVs (the deliverable, not
    just the inputs): row constants, canonical extension formulas, Summary
    SUMIF/SUM ranges, and the markup waterfall chain (labels, rates, bases)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        rep.add("WARN", "xlsx-tie-out", "openpyxl not installed — estimate.xlsx not verified")
        return
    wb = load_workbook(xlsx_path)
    if "Detail" not in wb.sheetnames or "Summary" not in wb.sheetnames:
        rep.add("FAIL", "xlsx-tie-out", "estimate.xlsx missing Detail/Summary sheet")
        return
    det, summ = wb["Detail"], wb["Summary"]
    rows = sorted(data, key=lambda r: (str(r.get("division", "")).zfill(3),
                                       str(r.get("section", ""))))
    n = len(rows)
    last = n + 1
    bad = []
    if det.max_row != last:
        bad.append(f"Detail has {det.max_row - 1} data rows, CSV has {n}")

    def close(a, b):
        return isinstance(a, (int, float)) and abs(a - b) <= 0.005

    for i, src in enumerate(rows):
        r = i + 2
        if r > det.max_row:
            break
        for col, key in ((5, "qty"), (7, "unit_mat"), (8, "waste_pct"), (10, "unit_lab"),
                         (12, "unit_equip"), (14, "unit_sub")):
            v = det.cell(r, col).value
            want = num(src[key])
            want = 0.0 if want is None else want
            if not close(v if v is not None else 0.0, want):
                bad.append(f"Detail!{det.cell(r, col).coordinate} = {v!r}, CSV says {want}")
        for col, want in ((9, f"=E{r}*G{r}*(1+H{r}/100)"), (11, f"=E{r}*J{r}"),
                          (13, f"=E{r}*L{r}"), (15, f"=E{r}*N{r}"),
                          (16, f"=I{r}+K{r}+M{r}+O{r}")):
            v = det.cell(r, col).value
            if v != want:
                bad.append(f"Detail!{det.cell(r, col).coordinate} formula {v!r} != canonical {want!r}")

    divs = sorted({str(x.get("division", "")).strip() for x in rows
                   if str(x.get("division", "")).strip()}, key=lambda x: x.zfill(3))
    first_div, direct_row = 7, 7 + len(divs)
    for i, dv in enumerate(divs):
        r = first_div + i
        if str(summ.cell(r, 1).value).strip() != dv:
            bad.append(f"Summary!A{r} = {summ.cell(r, 1).value!r}, expected div {dv!r}")
        for c, col in ((2, "I"), (3, "K"), (4, "M"), (5, "O")):
            want = f"=SUMIF(Detail!$A$2:$A${last},A{r},Detail!${col}$2:${col}${last})"
            if summ.cell(r, c).value != want:
                bad.append(f"Summary!{summ.cell(r, c).coordinate} formula "
                           f"{summ.cell(r, c).value!r} != {want!r}")
        if summ.cell(r, 6).value != f"=SUM(B{r}:E{r})":
            bad.append(f"Summary!F{r} formula {summ.cell(r, 6).value!r} != '=SUM(B{r}:E{r})'")
    for c, col in ((2, "B"), (3, "C"), (4, "D"), (5, "E"), (6, "F")):
        want = f"=SUM({col}{first_div}:{col}{direct_row - 1})"
        if summ.cell(direct_row, c).value != want:
            bad.append(f"Summary!{summ.cell(direct_row, c).coordinate} "
                       f"{summ.cell(direct_row, c).value!r} != {want!r} (division subtotal range)")

    w = direct_row + 4                       # waterfall "Direct Cost" line
    if summ.cell(w, 6).value != f"=F{direct_row}":
        bad.append(f"Summary!F{w} {summ.cell(w, 6).value!r} != '=F{direct_row}' (waterfall anchor)")
    seq = [("material_sales_tax_pct", "Material Sales Tax (materials only)", f"B{direct_row}")]
    seq += [(k, lbl, None) for k, lbl in MARKUP_ORDER]
    sub_row, r = w, w + 1
    for key, lbl, base in seq:
        base_ref = base or f"F{sub_row}"
        want_formula = f"={base_ref}*E{r}/100"
        want_rate = mk.get(key, 0.0)
        got_rate = summ.cell(r, 5).value
        if summ.cell(r, 1).value != lbl:
            bad.append(f"Summary!A{r} label {summ.cell(r, 1).value!r} != {lbl!r} (markup order)")
        if not close(got_rate if got_rate is not None else 0.0, want_rate):
            bad.append(f"Summary!E{r} rate {got_rate!r} != markups.csv {want_rate}")
        if summ.cell(r, 6).value != want_formula:
            bad.append(f"Summary!F{r} formula {summ.cell(r, 6).value!r} != {want_formula!r}")
        if summ.cell(r + 1, 6).value != f"=F{sub_row}+F{r}":
            bad.append(f"Summary!F{r + 1} subtotal {summ.cell(r + 1, 6).value!r} != '=F{sub_row}+F{r}'")
        sub_row, r = r + 1, r + 2
    if summ.cell(r, 1).value != "BID TOTAL" or summ.cell(r, 6).value != f"=F{sub_row}":
        bad.append(f"Summary!row {r}: expected BID TOTAL '=F{sub_row}', got "
                   f"{summ.cell(r, 1).value!r} / {summ.cell(r, 6).value!r}")

    if bad:
        rep.add("FAIL", "xlsx-tie-out",
                f"estimate.xlsx diverges from the CSVs / canonical formulas (rebuild or "
                f"reconcile): {bad if len(bad) <= 6 else bad[:6] + ['…+' + str(len(bad) - 6)]}")
    else:
        rep.add("PASS", "xlsx-tie-out",
                f"estimate.xlsx matches the CSVs ({n} rows) and the canonical waterfall chain")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("project_dir")
    ap.add_argument("--sector", default="residential",
                    choices=["residential", "commercial", "ti", "public"])
    ap.add_argument("--strict", action="store_true",
                    help="exit 1 on WARNs too (CI gate)")
    ap.add_argument("--xlsx", default=None,
                    help="path to estimate.xlsx (default: <project_dir>/estimate.xlsx)")
    args = ap.parse_args()
    proj = Path(args.project_dir)
    rep = Report()

    li_path = proj / "lineitems.csv"
    if not li_path.exists():
        print(f"FAIL: {li_path} not found"); sys.exit(1)
    with li_path.open(newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))

    # --- schema ---
    if rows[0] != HEADER:
        rep.add("FAIL", "schema", f"header mismatch: {rows[0]}")
    bad_width = [i + 1 for i, r in enumerate(rows) if len(r) != len(HEADER)]
    if bad_width:
        rep.add("FAIL", "schema", f"rows with wrong column count: {bad_width}")
    else:
        rep.add("PASS", "schema", f"{len(rows)-1} rows, 12 columns, header exact")

    data = [dict(zip(HEADER, r)) for r in rows[1:] if any(c.strip() for c in r)]

    # --- numeric parse + per-line checks ---
    tot = {"mat": 0.0, "lab": 0.0, "eq": 0.0, "sub": 0.0}
    by_div = defaultdict(float)
    allow_total, allow_count, ls_total, ls_count = 0.0, 0, 0.0, 0
    seen_items = defaultdict(list)
    zero_qty_bad, zero_qty_info, rollup_bad, waste_odd, mixed, bad_units, parse_bad = [], [], [], [], [], [], []
    bad_divs, line_exts = [], []

    for idx, r in enumerate(data, start=2):
        q = num(r["qty"]); w = num(r["waste_pct"])
        um, ul, ue, us = (num(r["unit_mat"]), num(r["unit_lab"]),
                          num(r["unit_equip"]), num(r["unit_sub"]))
        if None in (q, w, um, ul, ue, us):
            parse_bad.append(idx); continue
        unit = r["unit"].strip().upper()
        if unit and unit not in UNIT_WHITELIST:
            bad_units.append(f"row {idx}: '{r['unit']}'")
        label = f"row {idx} [{r['division']}] {r['item'][:44]}"
        div = r["division"].strip()
        if div not in CSI_DIVS:
            bad_divs.append(f"row {idx}: division={div!r}")
        is_rollup = bool(ROLLUP_RE.search(r["item"] + " " + r["description"]))
        costs = um + ul + ue + us

        if is_rollup and costs > 0:
            rollup_bad.append(label)
        if q == 0:
            if costs > 0:
                zero_qty_bad.append(label)  # priced placeholder that extends to $0
            elif not (INFO_NOTE_RE.search(r["notes"]) or ZERO_OK_NOTE_RE.search(r["notes"])
                      or is_rollup):
                zero_qty_info.append(label)
        if w and w > 0 and um == 0:
            waste_odd.append(label)
        if us > 0 and (um + ul + ue) > 0:
            mixed.append(label)

        mat = q * um * (1 + (w or 0) / 100)
        tot["mat"] += mat; tot["lab"] += q * ul; tot["eq"] += q * ue; tot["sub"] += q * us
        by_div[r["division"].strip()] += mat + q * (ul + ue + us)
        line_exts.append((div, idx, unit, mat + q * (ul + ue + us)))
        if unit == "ALLOW":
            allow_count += 1; allow_total += q * costs
        elif unit == "LS":
            ls_count += 1; ls_total += q * costs
        seen_items[(r["division"].strip(), r["item"].strip().lower())].append(idx)

    # qty-outlier screen: a non-LS/ALLOW line whose extension dwarfs the rest of
    # its division is usually a qty/unit typo (10x slip, SF-vs-SY, etc.).
    qty_outliers = []
    ext_by_div = defaultdict(list)
    for div, idx, unit, ext in line_exts:
        ext_by_div[div].append((idx, unit, ext))
    for div, entries in ext_by_div.items():
        if len(entries) < 2:
            continue
        for idx, unit, ext in entries:
            if unit in ("LS", "ALLOW") or ext <= 10000:
                continue
            med = median([e for i2, _u, e in entries if i2 != idx])
            if med > 0 and ext > 5 * med:
                qty_outliers.append(f"row {idx} [{div}] ext ${ext:,.0f} vs "
                                    f"${med:,.0f} median of the division's other lines")

    for check, bad, level, msg in [
        ("divisions", bad_divs, "FAIL",
         "blank/non-CSI division — these rows silently drop out of the Summary "
         "rollup and the BID TOTAL (use zero-padded 01-49)"),
        ("parse", parse_bad, "FAIL", "non-numeric qty/cost/waste at rows"),
        ("rollup-guard", rollup_bad, "FAIL", "rollup/total rows carrying unit costs (double-count risk)"),
        ("zero-qty", zero_qty_bad, "FAIL", "qty=0 with unit costs (unpriced commitment — price it or make it an allowance)"),
        ("zero-qty", zero_qty_info, "WARN", "qty=0 with no disposition note (mark rollup/info/RFI/by-others)"),
        ("waste-placement", waste_odd, "WARN", "waste_pct set but unit_mat=0 (waste applies to material only)"),
        ("mixed-line", mixed, "WARN", "line has both self-perform and sub costs"),
        ("units", bad_units, "WARN", "unit not in whitelist"),
        ("qty-outlier", qty_outliers, "WARN",
         "line extension > 5x its division's other-line median — verify qty/unit"),
    ]:
        if bad:
            rep.add(level, check, f"{msg}: {bad if len(bad) <= 8 else bad[:8] + ['…+' + str(len(bad)-8)]}")
        else:
            rep.add("PASS", check, "clean")

    dupes = {k: v for k, v in seen_items.items() if len(v) > 1}
    if dupes:
        rep.add("WARN", "duplicates", f"{len(dupes)} repeated (division,item) pairs — verify not double-counted: "
                f"{list(dupes.items())[:4]}")
    else:
        rep.add("PASS", "duplicates", "no repeated (division,item) pairs")

    direct = sum(tot.values())

    # --- markups + waterfall ---
    mk = {}
    mk_path = proj / "markups.csv"
    if mk_path.exists():
        with mk_path.open(newline="", encoding="utf-8-sig") as f:
            for r in csv.reader(f):
                if len(r) >= 2 and r[0].strip() and not r[0].strip().startswith("#"):
                    key = r[0].strip()
                    if key.lower() == "key":          # header row
                        continue
                    v = num(r[1])
                    if v is None:
                        rep.add("FAIL", "markups",
                                f"{key} value not numeric: {r[1]!r} — a builder would silently apply 0%")
                        v = 0.0
                    mk[key] = v
        for key, (lo, hi) in MARKUP_SANE.items():
            if key not in mk:
                rep.add("FAIL", "markups",
                        f"missing key {key} — builders diverge on defaults; set it explicitly (0 if none)")
            elif not (lo <= mk[key] <= hi):
                rep.add("WARN", "markups", f"{key}={mk[key]}% outside sane range {lo}-{hi}%")
        sub = direct + tot["mat"] * mk.get("material_sales_tax_pct", 0) / 100
        for key in ("general_conditions_pct", "contingency_pct", "insurance_pct",
                    "bond_pct", "permit_pct", "ohp_pct"):
            sub *= 1 + mk.get(key, 0) / 100
        rep.add("INFO", "waterfall",
                f"direct ${direct:,.0f} (mat ${tot['mat']:,.0f} incl waste) → BID TOTAL ${sub:,.0f}")
        if args.sector == "public" and mk.get("bond_pct", 0) == 0:
            rep.add("WARN", "markups", "sector=public but bond_pct=0 — FL 255.05 requires P&P bond on most public work")
    else:
        rep.add("WARN", "markups", "markups.csv not found — waterfall not recomputed")

    # --- benchmark bands ---
    bands = BANDS[args.sector]
    out_of_band = []
    for div, amt in sorted(by_div.items()):
        pct = 100 * amt / direct if direct else 0
        band = bands.get(div)
        if band and not (band[0] <= pct <= band[1]):
            out_of_band.append(f"Div {div}: {pct:.1f}% of direct (band {band[0]}-{band[1]}%)")
    if out_of_band:
        rep.add("WARN", "benchmark-bands", "; ".join(out_of_band) + " — justify in writing or re-price")
    else:
        rep.add("PASS", "benchmark-bands", f"all priced divisions inside {args.sector} bands")
    missing = EXPECTED_DIVS[args.sector] - {d for d, amt in by_div.items() if amt > 0}
    if missing:
        rep.add("WARN", "missing-divisions",
                f"expected for {args.sector} but $0/absent: {sorted(missing)} — price or explicitly exclude in scope")
    else:
        rep.add("PASS", "missing-divisions", "all expected divisions carry cost")

    # --- allowance tie-out vs scope (best effort) ---
    scope = proj / "scope-of-work.md"
    rep.add("INFO", "allowances", f"{allow_count} ALLOW rows ${allow_total:,.0f}; "
            f"{ls_count} LS (lump-sum work) rows ${ls_total:,.0f} — tie-out compares ALLOW only")
    if scope.exists():
        text = scope.read_text(errors="ignore")
        m = re.search(r"^#+.*allowance.*$", text, re.I | re.M)
        if m:
            section = text[m.end():]
            nxt = re.search(r"^#+\s", section, re.M)
            section = section[:nxt.start()] if nxt else section
            amts = [float(a.replace(",", "")) for a in re.findall(r"\$\s?([\d,]+(?:\.\d+)?)", section)]
            scope_total = sum(amts)
            if scope_total and abs(scope_total - allow_total) / max(scope_total, 1) > 0.02:
                rep.add("FAIL", "allowance-tie-out",
                        f"scope allowances ≈ ${scope_total:,.0f} vs CSV ALLOW ${allow_total:,.0f} "
                        f"(>2% apart) — reconcile before release (best-effort text scan; verify manually)")
            elif scope_total:
                rep.add("PASS", "allowance-tie-out",
                        f"scope ≈ ${scope_total:,.0f} vs CSV ${allow_total:,.0f} within 2% (best-effort scan)")
            else:
                rep.add("INFO", "allowance-tie-out", "no $ amounts parsed from scope allowance section")
        else:
            rep.add("INFO", "allowance-tie-out", "no Allowances heading found in scope-of-work.md")
    else:
        rep.add("INFO", "allowance-tie-out", "scope-of-work.md not present — tie-out skipped")

    # --- estimate.xlsx tie-out (the deliverable itself, not just the inputs) ---
    xlsx_path = Path(args.xlsx) if args.xlsx else proj / "estimate.xlsx"
    if xlsx_path.exists():
        tie_out_xlsx(xlsx_path, data, mk, rep)
    else:
        rep.add("INFO", "xlsx-tie-out", f"{xlsx_path.name} not present — workbook tie-out skipped")

    print(f"\nvalidate_estimate — {proj}  (sector={args.sector})\n")
    counts = rep.dump()
    sys.exit(1 if counts["FAIL"] or (args.strict and counts["WARN"]) else 0)


if __name__ == "__main__":
    main()
