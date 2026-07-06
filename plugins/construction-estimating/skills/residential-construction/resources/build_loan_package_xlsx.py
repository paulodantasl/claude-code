#!/usr/bin/env python3
"""Build a bank-ready Construction Loan Package workbook for a project.

Usage:
    python3 build_loan_package_xlsx.py <project_dir>
    (run the copy under estimating/scripts/ in the repo checkout, or
     ${CLAUDE_PLUGIN_ROOT}/scripts/ on a plugin install)

Reads:
    <project_dir>/lineitems.csv
    <project_dir>/markups.csv
    <project_dir>/loan-package-config.json
    <project_dir>/logo.png                       (optional; or company.logo in config)

Writes:
    <project_dir>/construction-loan-package.xlsx

Tabs:
    1. Cover
    2. Inputs                 (editable control panel — named ranges drive the workbook)
    3. Executive Summary
    4. Sources & Uses
    5. Budget Summary         (cost by CSI div + markup waterfall)
    6. Budget Detail          (every priced line item, formula-driven)
    7. Schedule of Values     (AIA G703-style)
    8. Draw Schedule          (monthly cash flow over the construction period)
    9. Construction Timeline  (Gantt-style by phase)
   10. Scope of Work          (narrative by CSI division)
   11. Allowances
   12. Alternates & Unit Prices
   13. Documents Checklist
"""

import csv
import json
import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


# ---------------------------------------------------------------------------
# Brand palette (matches the Ideal Construction logo: navy + gray)
NAVY = "1F3A6C"
NAVY_DK = "152849"
GRAY = "595D62"
LIGHT_BLUE = "DDE7F0"
LIGHT_GRAY = "F2F2F2"
ROW_ZEBRA = "F8F9FB"
ACCENT_GREEN = "2E7D32"
ACCENT_RED = "B91C1C"
WHITE = "FFFFFF"

CURRENCY = '_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-'
CURRENCY0 = '_-$* #,##0_-;-$* #,##0_-;_-$* "-"_-;_-@_-'
PCT1 = '0.0%'
PCT0 = '0%'
DATEFMT = 'mmm d, yyyy'

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
THICK = Side(style="medium", color=NAVY)
BORDER_THICK = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)


# ---------------------------------------------------------------------------
# Style helpers
def fill(rgb):
    return PatternFill("solid", fgColor=rgb)


def F_TITLE():
    return Font(name="Calibri", size=20, bold=True, color=NAVY)


def F_SUBTITLE():
    return Font(name="Calibri", size=12, italic=True, color=GRAY)


def F_H1():
    return Font(name="Calibri", size=14, bold=True, color=WHITE)


def F_H2():
    return Font(name="Calibri", size=11, bold=True, color=WHITE)


def F_LABEL():
    return Font(name="Calibri", size=10, bold=True, color=NAVY)


def F_BODY():
    return Font(name="Calibri", size=10, color="222222")


def F_TOTAL():
    return Font(name="Calibri", size=11, bold=True, color=NAVY)


def F_BIG_TOTAL():
    return Font(name="Calibri", size=14, bold=True, color=WHITE)


def F_NOTE():
    return Font(name="Calibri", size=9, italic=True, color=GRAY)


def header_row(ws, row, cols, fill_rgb=NAVY, font=None):
    font = font or F_H2()
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = fill(fill_rgb)
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def section_header(ws, row, text, span=12, fill_rgb=NAVY):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row, 1, text)
    cell.fill = fill(fill_rgb)
    cell.font = F_H1()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22


# ---------------------------------------------------------------------------
# Logo placement (skip first ~3 rows on every sheet for the brand strip)
def add_logo(ws, logo_path, anchor="A1"):
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    if not logo_path or not Path(logo_path).exists():
        return            # logo is optional — workbook still builds without it
    try:
        img = XLImage(str(logo_path))
    except ImportError:
        if not getattr(add_logo, "_warned_no_pillow", False):
            add_logo._warned_no_pillow = True
            print("Note: Pillow not installed — building without the logo image "
                  "(fix: pip install pillow).")
        return
    img.width = 220
    img.height = 70
    img.anchor = anchor
    ws.add_image(img)


def page_header(ws, title, subtitle, span=12):
    """Add 3-row brand strip with title + subtitle to the right of the logo."""
    ws.merge_cells(start_row=1, start_column=4, end_row=2, end_column=span)
    c = ws.cell(1, 4, title)
    c.font = F_TITLE()
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=span)
    c = ws.cell(3, 4, subtitle)
    c.font = F_SUBTITLE()
    c.alignment = Alignment(horizontal="right", vertical="center")


def define_name(wb, name, ref):
    """Define a workbook-level name. ref like 'Inputs!$C$10'."""
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names[name] = DefinedName(name, attr_text=ref)


def numfmt(cell, fmt):
    cell.number_format = fmt
    return cell


def setval(ws, r, c, v, font=None, fmt=None, align=None, fill_rgb=None, border=True):
    cell = ws.cell(r, c, v)
    if font:
        cell.font = font
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = align
    if fill_rgb:
        cell.fill = fill(fill_rgb)
    if border:
        cell.border = BORDER
    return cell


# ---------------------------------------------------------------------------
# Data loading
def _num(value, default=0.0):
    try:
        s = str(value).strip().replace("$", "").replace(",", "").replace("%", "")
        return float(s) if s else default
    except (TypeError, ValueError):
        return default


def load_lineitems(path):
    rows = []
    with path.open(newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            if not any((v or "").strip() for v in r.values()):
                continue
            rows.append(r)
    rows.sort(key=lambda r: (str(r.get("division", "")).zfill(3),
                             str(r.get("section", ""))))
    return rows


def load_markups(path):
    mk = {}
    if not path.exists():
        return mk
    with path.open(newline="", encoding="utf-8-sig") as f:
        for r in csv.reader(f):
            if len(r) >= 2 and r[0].strip() and not r[0].strip().startswith("#"):
                mk[r[0].strip()] = _num(r[1])
    return mk


def load_config(path):
    return json.loads(path.read_text())


# ---------------------------------------------------------------------------
# Division names (CSI MasterFormat 2018)
DIV_NAMES = {
    "01": "General Requirements",
    "02": "Existing Conditions",
    "03": "Concrete",
    "04": "Masonry",
    "05": "Metals",
    "06": "Wood, Plastics & Composites",
    "07": "Thermal & Moisture Protection",
    "08": "Openings",
    "09": "Finishes",
    "10": "Specialties",
    "11": "Equipment",
    "12": "Furnishings",
    "13": "Special Construction",
    "14": "Conveying Equipment",
    "21": "Fire Suppression",
    "22": "Plumbing",
    "23": "HVAC",
    "25": "Integrated Automation",
    "26": "Electrical",
    "27": "Communications",
    "28": "Electronic Safety & Security",
    "31": "Earthwork",
    "32": "Exterior Improvements",
    "33": "Utilities",
}


# Typical residential phasing matrix (% per month for an 11-month build).
# Each list must sum to 100 and have N entries equal to schedule.duration_months.
DEFAULT_PHASING_11M = {
    "01": [10, 9, 9, 9, 9, 9, 9, 9, 9, 9, 9],   # GCs evenly spread (only if Div 01 explicit items)
    "02": [50, 50, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    "03": [10, 30, 30, 20, 10, 0, 0, 0, 0, 0, 0],
    "04": [0, 0, 10, 35, 40, 15, 0, 0, 0, 0, 0],
    "05": [0, 0, 0, 15, 35, 30, 15, 5, 0, 0, 0],
    "06": [0, 0, 0, 10, 35, 35, 15, 5, 0, 0, 0],
    "07": [0, 0, 5, 5, 15, 35, 25, 10, 5, 0, 0],
    "08": [0, 0, 0, 0, 0, 25, 40, 25, 10, 0, 0],
    "09": [0, 0, 0, 0, 0, 5, 20, 35, 25, 12, 3],
    "10": [0, 0, 0, 0, 0, 0, 0, 25, 50, 20, 5],
    "11": [0, 0, 0, 0, 0, 0, 0, 0, 30, 60, 10],
    "12": [0, 0, 0, 0, 0, 0, 0, 0, 30, 60, 10],
    "13": [0, 0, 0, 0, 25, 35, 30, 10, 0, 0, 0],
    "14": [0, 0, 0, 0, 0, 0, 20, 40, 40, 0, 0],
    "21": [0, 0, 0, 0, 0, 0, 50, 50, 0, 0, 0],
    "22": [0, 0, 5, 10, 20, 25, 20, 10, 5, 3, 2],
    "23": [0, 0, 0, 5, 10, 20, 25, 20, 10, 5, 5],
    "26": [3, 3, 5, 8, 12, 18, 20, 15, 8, 5, 3],
    "27": [0, 0, 0, 0, 0, 10, 30, 40, 15, 5, 0],
    "28": [0, 0, 0, 0, 0, 0, 15, 40, 30, 10, 5],
    "31": [25, 30, 25, 10, 0, 0, 0, 0, 0, 5, 5],
    "32": [0, 0, 0, 0, 0, 0, 0, 5, 10, 40, 45],
    "33": [0, 20, 30, 30, 20, 0, 0, 0, 0, 0, 0],
}


def phasing_for_n(months):
    """Return phasing matrix scaled to N months (linear stretch)."""
    if months == 11:
        return DEFAULT_PHASING_11M
    # simple resample: for each division, linearly interpolate the 11-month curve
    out = {}
    n = months
    for div, curve11 in DEFAULT_PHASING_11M.items():
        if n == 11:
            out[div] = curve11
            continue
        # resample by linear interpolation
        new = [0.0] * n
        for i in range(n):
            src = i * (10) / (n - 1) if n > 1 else 0
            lo, hi = int(src), min(int(src) + 1, 10)
            frac = src - lo
            new[i] = curve11[lo] * (1 - frac) + curve11[hi] * frac
        # normalize to 100
        s = sum(new)
        if s > 0:
            new = [v * 100 / s for v in new]
        out[div] = new
    return out


# ---------------------------------------------------------------------------
# BUILD TABS
def build_cover(wb, cfg, logo):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    add_logo(ws, logo, "A1")

    # Title block
    ws.merge_cells("B6:H7")
    cell = ws.cell(6, 2, "CONSTRUCTION LOAN PACKAGE")
    cell.font = Font(name="Calibri", size=28, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B9:H9")
    cell = ws.cell(9, 2, cfg["project"]["name"])
    cell.font = Font(name="Calibri", size=18, bold=True, color=GRAY)
    cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("B10:H10")
    cell = ws.cell(10, 2, f'{cfg["project"]["address"]}, {cfg["project"]["city_state_zip"]}')
    cell.font = Font(name="Calibri", size=12, italic=True, color=GRAY)
    cell.alignment = Alignment(horizontal="center")

    # GC presenter box (left)
    ws.merge_cells("B14:D14"); ws["B14"] = "Prepared & Submitted By"
    ws["B14"].font = F_H2(); ws["B14"].fill = fill(NAVY); ws["B14"].alignment = Alignment(horizontal="center")
    rows = [
        (cfg["company"]["name"], True),
        (cfg["company"].get("license", ""), False),
        (cfg["company"]["address"], False),
        (cfg["company"]["city_state_zip"], False),
        (cfg["company"]["phone"], False),
        (cfg["company"]["email"], False),
    ]
    for i, (val, bold) in enumerate(rows):
        ws.merge_cells(start_row=15 + i, start_column=2, end_row=15 + i, end_column=4)
        c = ws.cell(15 + i, 2, val)
        c.font = Font(name="Calibri", size=11, bold=bold, color="222222")
        c.alignment = Alignment(horizontal="center")

    # Lender / Borrower box (right)
    ws.merge_cells("F14:H14"); ws["F14"] = "Prepared For"
    ws["F14"].font = F_H2(); ws["F14"].fill = fill(NAVY); ws["F14"].alignment = Alignment(horizontal="center")
    rows = [
        ("=Inputs!C20", True),         # lender name
        ("=Inputs!C21", False),        # lender contact
        ("=Inputs!C22", False),        # loan #
        ("", False),
        ("Borrower:", False),
        ("=Inputs!C16", True),         # owner name
    ]
    for i, (val, bold) in enumerate(rows):
        ws.merge_cells(start_row=15 + i, start_column=6, end_row=15 + i, end_column=8)
        c = ws.cell(15 + i, 6, val)
        c.font = Font(name="Calibri", size=11, bold=bold, color="222222")
        c.alignment = Alignment(horizontal="center")

    # Big number — BID TOTAL
    ws.merge_cells("B24:H24")
    c = ws.cell(24, 2, "TOTAL PROJECT COST")
    c.font = F_H2(); c.fill = fill(NAVY); c.alignment = Alignment(horizontal="center")

    ws.merge_cells("B25:H26")
    c = ws.cell(25, 2, "=Sources_and_Uses!C16")  # Total Project Cost
    c.font = Font(name="Calibri", size=36, bold=True, color=NAVY)
    c.number_format = CURRENCY0
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B28:H28")
    c = ws.cell(28, 2, "=CONCATENATE(\"$\", TEXT(Bid_Total/Inputs!C6, \"#,##0\"), \" per SF conditioned   |   \", "
                       "TEXT(Inputs!C6, \"#,##0\"), \" SF conditioned   |   \", "
                       "TEXT(Inputs!C7, \"#,##0\"), \" SF gross under-roof\")")
    c.font = F_SUBTITLE(); c.alignment = Alignment(horizontal="center")

    # Date / revision
    ws.merge_cells("B31:H31")
    c = ws.cell(31, 2, f'Issued: {date.today().strftime("%B %d, %Y")}  |  Revision 0  |  Confidential')
    c.font = F_NOTE(); c.alignment = Alignment(horizontal="center")

    # Document index
    ws.merge_cells("B34:H34"); ws["B34"] = "Document Index"
    ws["B34"].font = F_H2(); ws["B34"].fill = fill(NAVY); ws["B34"].alignment = Alignment(horizontal="center")
    index = [
        ("1.  Inputs / Project Data", "Editable control panel — drives all calculated tabs"),
        ("2.  Executive Summary", "Project, code/wind/flood, schedule, financial summary, key contacts"),
        ("3.  Sources & Uses", "Construction loan + owner equity vs. hard + soft costs + contingency"),
        ("4.  Budget Summary", "Direct cost by CSI division + full markup waterfall"),
        ("5.  Budget Detail", "Every priced line item with live formulas (Detail-level)"),
        ("6.  Schedule of Values", "AIA G703-style scheduled values for draw requests"),
        ("7.  Draw Schedule", "Monthly cash-flow forecast over the construction period"),
        ("8.  Construction Timeline", "Gantt-style phasing"),
        ("9.  Scope of Work", "Inclusions / exclusions / clarifications by CSI division"),
        ("10. Allowances", "All allowance items with $ values and basis"),
        ("11. Alternates & Unit Prices", "Owner / lender selection items"),
        ("12. Documents Checklist", "Licensing, insurance, bonds, geotech, permits, certificates"),
    ]
    for i, (label, desc) in enumerate(index):
        ws.cell(36 + i, 2, label).font = Font(name="Calibri", size=10, bold=True, color=NAVY)
        ws.merge_cells(start_row=36 + i, start_column=4, end_row=36 + i, end_column=8)
        ws.cell(36 + i, 4, desc).font = F_BODY()

    # Column widths
    for i, w in enumerate([2, 18, 18, 18, 18, 18, 18, 18, 4], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.zoomScale = 100


def build_inputs(wb, cfg, logo):
    ws = wb.create_sheet("Inputs")
    ws.sheet_view.showGridLines = False
    add_logo(ws, logo)
    page_header(ws, "Inputs / Project Data", "Edit the green cells. All other tabs update automatically.")

    # Helper to render a labelled input row with a defined name
    def row(r, label, value, name=None, fmt=None, comment=None):
        ws.cell(r, 2, label).font = F_LABEL()
        ws.cell(r, 2).alignment = Alignment(horizontal="right", indent=1)
        c = ws.cell(r, 3, value)
        c.fill = fill("DDF7E0")           # editable cells = light green
        c.font = F_BODY()
        c.alignment = Alignment(horizontal="left", indent=1)
        c.border = BORDER
        if fmt:
            c.number_format = fmt
        if name:
            define_name(wb, name, f"Inputs!$C${r}")
        if comment:
            ws.cell(r, 4, comment).font = F_NOTE()
        return c

    # Section: Project
    section_header(ws, 5, "PROJECT", span=10)
    row(6, "Conditioned area (SF)", cfg["project"]["conditioned_sf"], "Conditioned_SF", "#,##0")
    row(7, "Gross under-roof (SF)", cfg["project"]["gross_under_roof_sf"], "Gross_SF", "#,##0")
    row(8, "Lot area (SF)", cfg["project"]["lot_sf"], "Lot_SF", "#,##0")
    row(9, "Project name", cfg["project"]["name"], "Project_Name")
    row(10, "Address", cfg["project"]["address"], "Project_Address")
    row(11, "City, State, Zip", cfg["project"]["city_state_zip"], "Project_City")
    row(12, "County / AHJ", cfg["project"]["county"], "Project_County")
    row(13, "Occupancy / Construction Type", f'{cfg["project"]["occupancy"]} / {cfg["project"]["construction_type"]}', "Occupancy_Type")
    row(14, "Wind / Flood", f'V_ult {cfg["project"]["vult_mph"]} mph Exp {cfg["project"]["exposure"]}; Zone {cfg["project"]["flood_zone"]}; BFE {cfg["project"]["bfe_ft"]} ft', "Wind_Flood")

    # Section: Owner & Team
    section_header(ws, 16, "OWNER & PROJECT TEAM", span=10, fill_rgb=GRAY)
    row(17, "Owner / Borrower", cfg["owner"]["name"], "Owner_Name")
    row(18, "Architect", cfg["architect"]["name"], "Architect_Name")
    row(19, "Structural Engineer", f'{cfg["engineer"]["name"]} ({cfg["engineer"]["principal"]})', "Engineer_Name")
    row(20, "Geotechnical Engineer", f'{cfg["geotech"]["firm"]} (#{cfg["geotech"]["report_no"]})', "Geotech_Name")

    # Section: Lender
    section_header(ws, 22, "LENDER", span=10, fill_rgb=GRAY)
    row(23, "Lender name", cfg["lender"]["name"], "Lender_Name")
    row(24, "Lender contact", cfg["lender"]["contact"], "Lender_Contact")
    row(25, "Loan number", cfg["lender"]["loan_no"], "Loan_No")

    # Section: GC
    section_header(ws, 27, "GENERAL CONTRACTOR", span=10, fill_rgb=GRAY)
    row(28, "GC company", cfg["company"]["name"], "GC_Name")
    row(29, "GC license", cfg["company"].get("license", ""), "GC_License")
    row(30, "GC address", cfg["company"]["address"], "GC_Address")
    row(31, "GC phone", cfg["company"]["phone"], "GC_Phone")
    row(32, "GC email", cfg["company"]["email"], "GC_Email")

    # Section: Schedule
    section_header(ws, 34, "SCHEDULE", span=10)
    row(35, "Target start date", date.fromisoformat(cfg["schedule"]["target_start_date"]), "Start_Date", DATEFMT)
    row(36, "Duration (months)", cfg["schedule"]["duration_months"], "Duration_Months", "0")
    row(37, "Retainage %", cfg["schedule"]["retainage_pct"] / 100, "Retainage_Pct", PCT1)

    # Section: Markups (linked to markups.csv loaded values, but user can override here)
    section_header(ws, 39, "MARKUPS & WATERFALL", span=10)
    row(40, "Material sales tax %", "=Material_Tax_Pct_Source", "Material_Tax_Pct", PCT1,
        "FL 6% + Pinellas 1%")
    row(41, "General Conditions %", "=GC_Pct_Source", "GC_Markup_Pct", PCT1)
    row(42, "Contingency %", "=Cont_Pct_Source", "Contingency_Pct", PCT1)
    row(43, "Insurance %", "=Ins_Pct_Source", "Insurance_Pct", PCT1)
    row(44, "Bond %", "=Bond_Pct_Source", "Bond_Pct", PCT1)
    row(45, "Permit %", "=Permit_Pct_Source", "Permit_Pct", PCT1)
    row(46, "OH&P %", "=OHP_Pct_Source", "OHP_Pct", PCT1)

    # Section: Loan financials
    section_header(ws, 48, "LOAN STRUCTURE", span=10)
    la = cfg["loan_assumptions"]
    row(49, "Land cost (if loan-funded)", la["land_cost"], "Land_Cost", CURRENCY0)
    row(50, "Soft costs % of hard cost", la["soft_costs_pct_of_hard"] / 100, "Soft_Costs_Pct", PCT1,
        "A&E + permits + survey + geotech + closing + interest reserve")
    row(51, "Owner soft contingency %", la["owner_soft_contingency_pct_of_hard"] / 100, "Owner_Contingency_Pct", PCT1,
        "In addition to the GC's contingency in the bid")
    row(52, "Owner equity %", la["owner_equity_pct"] / 100, "Owner_Equity_Pct", PCT1,
        "Lender typically requires 20-25% owner equity")
    row(53, "Interest reserve % of loan", la["interest_reserve_pct_of_loan"] / 100, "Interest_Reserve_Pct", PCT1)
    row(54, "Lender fees % of loan", la["lender_fees_pct_of_loan"] / 100, "Lender_Fees_Pct", PCT1)

    # Notes & how-to
    section_header(ws, 57, "HOW THIS WORKBOOK IS WIRED", span=10, fill_rgb=GRAY)
    notes = [
        "✎  Green cells are editable. All other tabs update via Excel formulas (named ranges).",
        "•  Change a MARKUP % above → Budget Summary waterfall + Sources & Uses + SOV + Draws all recompute.",
        "•  Change a UNIT COST on Budget Detail → the line total, division subtotal, BID TOTAL, and SOV/Draw cascade.",
        "•  Change DURATION (months) → Draw Schedule and Construction Timeline column count adjusts on next workbook regen.",
        "•  Change OWNER EQUITY % → Construction Loan amount and LTC update.",
        "•  SOV is AIA G702/G703-style: enter \"Work This Period\" + \"Materials Stored\" on green cells → retainage and pay-this-period compute.",
        "•  Sales tax base = material extensions only (column I on Detail), per FL convention.",
        "•  Subs are not re-burdened; OH&P applies once at the end of the waterfall.",
    ]
    for i, txt in enumerate(notes):
        ws.merge_cells(start_row=58 + i, start_column=2, end_row=58 + i, end_column=10)
        c = ws.cell(58 + i, 2, txt)
        c.font = F_BODY()
        c.alignment = Alignment(wrap_text=True, vertical="center", indent=1)

    # Column widths
    widths = [2, 32, 22, 32, 4]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"


def build_executive_summary(wb, cfg, logo):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    add_logo(ws, logo)
    page_header(ws, "Executive Summary", "655 115th Ave, Treasure Island, FL")

    section_header(ws, 5, "PROJECT OVERVIEW", span=10)
    overview = [
        ("Project", "=Project_Name"),
        ("Address", '=CONCATENATE(Project_Address, ", ", Project_City)'),
        ("County / AHJ", "=Project_County"),
        ("Use & Construction Type", "=Occupancy_Type"),
        ("Wind / Flood basis", "=Wind_Flood"),
        ("Conditioned SF / Gross under-roof", '=CONCATENATE(TEXT(Conditioned_SF,"#,##0")," / ",TEXT(Gross_SF,"#,##0"))'),
        ("Lot area (SF)", "=Lot_SF"),
        ("Code basis", cfg["project"]["fbc_edition"]),
    ]
    for i, (label, val) in enumerate(overview):
        setval(ws, 6 + i, 2, label, font=F_LABEL(), align=Alignment(horizontal="right"))
        ws.merge_cells(start_row=6 + i, start_column=3, end_row=6 + i, end_column=8)
        c = ws.cell(6 + i, 3, val)
        c.font = F_BODY()
        c.alignment = Alignment(horizontal="left", indent=1)
        c.border = BORDER

    section_header(ws, 16, "PROJECT TEAM", span=10, fill_rgb=GRAY)
    team = [
        ("General Contractor", "=GC_Name", "=GC_License"),
        ("Architect of Record", "=Architect_Name", ""),
        ("Structural Engineer", "=Engineer_Name", f'Project #{cfg["engineer"]["project_no"]}'),
        ("Geotechnical Engineer", "=Geotech_Name", ""),
        ("Owner / Borrower", "=Owner_Name", ""),
        ("Lender", "=Lender_Name", "=Loan_No"),
    ]
    for i, (label, name, lic) in enumerate(team):
        setval(ws, 17 + i, 2, label, font=F_LABEL(), align=Alignment(horizontal="right"))
        ws.merge_cells(start_row=17 + i, start_column=3, end_row=17 + i, end_column=5)
        c = ws.cell(17 + i, 3, name)
        c.font = F_BODY(); c.alignment = Alignment(horizontal="left", indent=1); c.border = BORDER
        ws.merge_cells(start_row=17 + i, start_column=6, end_row=17 + i, end_column=8)
        c = ws.cell(17 + i, 6, lic)
        c.font = F_NOTE(); c.alignment = Alignment(horizontal="left", indent=1); c.border = BORDER

    section_header(ws, 25, "FINANCIAL SUMMARY", span=10)
    fin = [
        ("Hard Construction Cost (Bid Total)", "=Bid_Total"),
        ("$ per SF conditioned", "=Bid_Total/Conditioned_SF"),
        ("$ per SF gross under-roof", "=Bid_Total/Gross_SF"),
        ("Soft Costs (per Inputs)", "=Sources_and_Uses!C13"),
        ("Owner Soft Contingency", "=Sources_and_Uses!C14"),
        ("Interest Reserve", "=Sources_and_Uses!C15"),
        ("TOTAL PROJECT COST", "=Sources_and_Uses!C16"),
        ("", ""),
        ("Owner Equity (per Inputs)", "=Sources_and_Uses!C20"),
        ("Construction Loan Requested", "=Sources_and_Uses!C21"),
        ("Loan-to-Cost ratio", "=Sources_and_Uses!C21/Sources_and_Uses!C16"),
    ]
    for i, (label, val) in enumerate(fin):
        is_total = "TOTAL" in label or "Loan" in label
        setval(ws, 26 + i, 2, label,
               font=F_TOTAL() if is_total else F_LABEL(),
               align=Alignment(horizontal="right"))
        ws.merge_cells(start_row=26 + i, start_column=3, end_row=26 + i, end_column=5)
        c = ws.cell(26 + i, 3, val)
        if "ratio" in label.lower() or "SF" in label.split(",")[0]:
            c.number_format = PCT1 if "ratio" in label.lower() else '$#,##0.00'
        else:
            c.number_format = CURRENCY0
        c.font = F_TOTAL() if is_total else F_BODY()
        if is_total:
            c.fill = fill(LIGHT_BLUE)
        c.alignment = Alignment(horizontal="right", indent=1)
        c.border = BORDER

    section_header(ws, 38, "SCHEDULE SUMMARY", span=10, fill_rgb=GRAY)
    sched = [
        ("Target Start Date", "=Start_Date", DATEFMT),
        ("Duration", '=CONCATENATE(Duration_Months," months")', None),
        ("Target Substantial Completion", "=EDATE(Start_Date, Duration_Months)", DATEFMT),
        ("Draw Frequency", "Monthly (per AIA G702/G703)", None),
        ("Retainage", "=Retainage_Pct", PCT1),
    ]
    for i, (label, val, fmt) in enumerate(sched):
        setval(ws, 39 + i, 2, label, font=F_LABEL(), align=Alignment(horizontal="right"))
        ws.merge_cells(start_row=39 + i, start_column=3, end_row=39 + i, end_column=5)
        c = ws.cell(39 + i, 3, val)
        if fmt:
            c.number_format = fmt
        c.font = F_BODY()
        c.alignment = Alignment(horizontal="left", indent=1)
        c.border = BORDER

    section_header(ws, 45, "KEY FLORIDA CODE FEATURES (FBC 2023)", span=10, fill_rgb=GRAY)
    fl = [
        "• Coastal Pinellas (NOT HVHZ — Miami-Dade/Broward only)",
        f'• Wind design V_ult {cfg["project"]["vult_mph"]} mph, Exposure {cfg["project"]["exposure"]}, Risk Cat II per ASCE 7',
        f'• Flood Zone {cfg["project"]["flood_zone"]}, BFE {cfg["project"]["bfe_ft"]} ft (NGVD 1988) — structure elevated on driven timber piles',
        "• Garage / open porch is non-habitable below DFE (FBC flood-vented, equipment-restricted)",
        "• Impact-resistant windows/doors per Windborne Debris Region (FBC 1609.1.2)",
        "• Subterranean termite soil treatment (FBC 1816)",
        "• Pre-engineered wood trusses (delegated design), hurricane-strapped (continuous load path)",
        "• Threshold inspection allowance carried (3-story / 47 ft borderline FL 553.79)",
    ]
    for i, t in enumerate(fl):
        ws.merge_cells(start_row=46 + i, start_column=2, end_row=46 + i, end_column=10)
        c = ws.cell(46 + i, 2, t)
        c.font = F_BODY()
        c.alignment = Alignment(horizontal="left", indent=1)

    widths = [2, 32, 22, 18, 18, 18, 14, 14, 4]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_sources_uses(wb, cfg, logo):
    ws = wb.create_sheet("Sources_and_Uses")
    ws.sheet_view.showGridLines = False
    add_logo(ws, logo)
    page_header(ws, "Sources & Uses", "Construction loan structure", span=8)

    section_header(ws, 5, "USES OF FUNDS", span=8, fill_rgb=NAVY)
    setval(ws, 6, 2, "Category", font=F_H2(), fill_rgb=GRAY, align=Alignment(horizontal="center"))
    setval(ws, 6, 3, "Amount", font=F_H2(), fill_rgb=GRAY, align=Alignment(horizontal="center"))
    setval(ws, 6, 4, "Basis / Notes", font=F_H2(), fill_rgb=GRAY, align=Alignment(horizontal="center"))

    uses_rows = [
        (7,  "Land Acquisition",                "=Land_Cost",                                              "Lot owned outright? Enter $0 if not loan-funded"),
        (8,  "Hard Construction Cost (Bid)",    "=Bid_Total",                                     "From Budget Summary (auto-linked)"),
        (9,  "  — Architectural & Engineering", "=C8*0.04",                                                "~4% A&E"),
        (10, "  — Survey, Geotech, Soils",      "=C8*0.005",                                               "Geotech + survey"),
        (11, "  — Permit / Impact fees (outside GC)", "=C8*0.015",                                         "Above-and-beyond GC permit allowance"),
        (12, "  — Legal / Title / Closing",     "=C8*0.005",                                               "Closing costs"),
        (13, "Soft Costs Subtotal",             "=SUM(C9:C12)",                                            "Sum of A&E, survey, permits, legal"),
        (14, "Owner Soft Contingency",          "=C8*Owner_Contingency_Pct",                               "Additional contingency above GC bid"),
        (15, "Interest Reserve",                "=C8*Interest_Reserve_Pct",                                "Budgetary ~3-5% of hard cost (breaks circular ref vs loan)"),
        (16, "TOTAL PROJECT COST",              "=C7+C8+C13+C14+C15",                                      "Land + Hard + Soft + Owner Cont. + Interest"),
    ]
    for r, label, val, note in uses_rows:
        is_total = "TOTAL" in label or label == "Soft Costs Subtotal"
        bold = is_total or label == "Hard Construction Cost (Bid)"
        setval(ws, r, 2, label,
               font=F_TOTAL() if bold else F_BODY(),
               align=Alignment(horizontal="left", indent=2),
               fill_rgb=LIGHT_BLUE if is_total else None)
        c = setval(ws, r, 3, val, fmt=CURRENCY0,
                   font=F_TOTAL() if bold else F_BODY(),
                   align=Alignment(horizontal="right"),
                   fill_rgb=LIGHT_BLUE if is_total else None)
        setval(ws, r, 4, note, font=F_NOTE(), align=Alignment(horizontal="left", indent=1))

    section_header(ws, 18, "SOURCES OF FUNDS", span=8, fill_rgb=NAVY)
    setval(ws, 19, 2, "Source", font=F_H2(), fill_rgb=GRAY, align=Alignment(horizontal="center"))
    setval(ws, 19, 3, "Amount", font=F_H2(), fill_rgb=GRAY, align=Alignment(horizontal="center"))
    setval(ws, 19, 4, "Basis / Notes", font=F_H2(), fill_rgb=GRAY, align=Alignment(horizontal="center"))

    sources_rows = [
        (20, "Owner Equity (Cash)",       "=C16*Owner_Equity_Pct",      f'Per Inputs (typically 20-25%)'),
        (21, "Construction Loan",         "=C16-C20",                   "Balance funded by lender"),
        (22, "TOTAL SOURCES",             "=C20+C21",                   "Must equal Total Uses"),
    ]
    for r, label, val, note in sources_rows:
        bold = "TOTAL" in label or "Loan" in label
        setval(ws, r, 2, label,
               font=F_TOTAL() if bold else F_BODY(),
               align=Alignment(horizontal="left", indent=2),
               fill_rgb=LIGHT_BLUE if bold else None)
        c = setval(ws, r, 3, val, fmt=CURRENCY0,
                   font=F_TOTAL() if bold else F_BODY(),
                   align=Alignment(horizontal="right"),
                   fill_rgb=LIGHT_BLUE if bold else None)
        setval(ws, r, 4, note, font=F_NOTE(), align=Alignment(horizontal="left", indent=1))

    # Sources-Uses balance check
    section_header(ws, 24, "BALANCE CHECK", span=8, fill_rgb=GRAY)
    setval(ws, 25, 2, "Total Uses", font=F_LABEL(), align=Alignment(horizontal="right"))
    setval(ws, 25, 3, "=C16", fmt=CURRENCY0, font=F_TOTAL(), align=Alignment(horizontal="right"))
    setval(ws, 26, 2, "Total Sources", font=F_LABEL(), align=Alignment(horizontal="right"))
    setval(ws, 26, 3, "=C22", fmt=CURRENCY0, font=F_TOTAL(), align=Alignment(horizontal="right"))
    setval(ws, 27, 2, "Difference (must be $0)", font=F_LABEL(), align=Alignment(horizontal="right"))
    c = setval(ws, 27, 3, "=C26-C25", fmt=CURRENCY, font=F_TOTAL(), align=Alignment(horizontal="right"))
    c.fill = fill(LIGHT_GRAY)

    # Lender quick metrics
    section_header(ws, 30, "LENDER METRICS", span=8, fill_rgb=GRAY)
    metrics = [
        ("Loan-to-Cost (LTC)",      "=C21/C16", PCT1, "Construction Loan / Total Project Cost"),
        ("Hard Cost / Total Cost",  "=C8/C16",  PCT1, "Hard cost as % of total project"),
        ("Soft Cost / Hard Cost",   "=C13/C8",  PCT1, "Soft cost as % of hard cost (15-20% normal range)"),
        ("Lender Fees (estimate)",  "=C21*Lender_Fees_Pct", CURRENCY0, "Origination + processing"),
    ]
    for i, (label, val, fmt, note) in enumerate(metrics):
        setval(ws, 31 + i, 2, label, font=F_LABEL(), align=Alignment(horizontal="right"))
        setval(ws, 31 + i, 3, val, fmt=fmt, font=F_BODY(), align=Alignment(horizontal="right"))
        setval(ws, 31 + i, 4, note, font=F_NOTE(), align=Alignment(horizontal="left", indent=1))

    widths = [2, 38, 22, 60, 4]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"


def build_budget_detail(wb, items, logo):
    ws = wb.create_sheet("Budget_Detail")
    add_logo(ws, logo)
    page_header(ws, "Budget Detail", "Every priced line item — formula-driven", span=18)

    headers = ["Div", "Section", "Item", "Description", "Qty", "Unit",
               "Unit Mat", "Waste %", "Mat Ext", "Unit Lab", "Lab Ext",
               "Unit Equip", "Equip Ext", "Unit Sub", "Sub Ext", "Line Total", "Notes"]
    ws.append([])  # blank row 4
    ws.append([])
    header_row_idx = 6
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(header_row_idx, c, h)
        cell.fill = fill(NAVY); cell.font = F_H2()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    start_data = header_row_idx + 1
    for i, it in enumerate(items):
        r = start_data + i
        ws.cell(r, 1, str(it.get("division", "")).strip())
        ws.cell(r, 2, str(it.get("section", "")).strip())
        ws.cell(r, 3, it.get("item", ""))
        ws.cell(r, 4, it.get("description", ""))
        ws.cell(r, 5, _num(it.get("qty")))
        ws.cell(r, 6, str(it.get("unit", "")).strip())
        ws.cell(r, 7, _num(it.get("unit_mat")))
        ws.cell(r, 8, _num(it.get("waste_pct")))
        ws.cell(r, 9, f"=E{r}*G{r}*(1+H{r}/100)")     # Mat Ext (pre-tax)
        ws.cell(r, 10, _num(it.get("unit_lab")))
        ws.cell(r, 11, f"=E{r}*J{r}")                  # Lab Ext
        ws.cell(r, 12, _num(it.get("unit_equip")))
        ws.cell(r, 13, f"=E{r}*L{r}")                  # Equip Ext
        ws.cell(r, 14, _num(it.get("unit_sub")))
        ws.cell(r, 15, f"=E{r}*N{r}")                  # Sub Ext
        ws.cell(r, 16, f"=I{r}+K{r}+M{r}+O{r}")        # Line Total
        ws.cell(r, 17, it.get("notes", ""))
        for c in (5, 7, 9, 10, 11, 12, 13, 14, 15, 16):
            ws.cell(r, c).number_format = CURRENCY if c != 5 else '#,##0.00'
        ws.cell(r, 8).number_format = '0.0"%"'
        for c in range(1, 18):
            ws.cell(r, c).border = BORDER
            if i % 2 == 1:
                ws.cell(r, c).fill = fill(ROW_ZEBRA)
        for c in (4, 17):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="center")

    last = ws.max_row
    widths = [6, 11, 22, 40, 9, 7, 11, 8, 13, 11, 13, 11, 13, 11, 13, 13, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{header_row_idx + 1}"
    # Define a range for the Summary to reference
    define_name(wb, "Detail_Div", f"Budget_Detail!$A${start_data}:$A${last}")
    define_name(wb, "Detail_Mat", f"Budget_Detail!$I${start_data}:$I${last}")
    define_name(wb, "Detail_Lab", f"Budget_Detail!$K${start_data}:$K${last}")
    define_name(wb, "Detail_Equip", f"Budget_Detail!$M${start_data}:$M${last}")
    define_name(wb, "Detail_Sub", f"Budget_Detail!$O${start_data}:$O${last}")
    define_name(wb, "Detail_Total", f"Budget_Detail!$P${start_data}:$P${last}")
    return start_data, last


def build_budget_summary(wb, items, markups, logo, detail_start, detail_last):
    ws = wb.create_sheet("Budget_Summary")
    ws.sheet_view.showGridLines = False
    add_logo(ws, logo)
    page_header(ws, "Budget Summary", "Direct cost by CSI division + full markup waterfall", span=8)

    # Markup source cells (linked to Inputs' named-range values; but for editability we
    # also keep them in a small "Markup Source" panel here for transparency)
    # Pre-stage the named ranges that Inputs' formulas reference back to here
    section_header(ws, 5, "MARKUP SOURCES (mirror of Inputs)", span=8, fill_rgb=GRAY)
    markup_panel = [
        ("Material sales tax %",       "Material_Tax_Pct_Source", markups.get("material_sales_tax_pct", 7.0) / 100),
        ("General Conditions %",       "GC_Pct_Source",            markups.get("general_conditions_pct", 10) / 100),
        ("Contingency %",              "Cont_Pct_Source",          markups.get("contingency_pct", 5) / 100),
        ("Insurance %",                "Ins_Pct_Source",           markups.get("insurance_pct", 1.2) / 100),
        ("Bond %",                     "Bond_Pct_Source",          markups.get("bond_pct", 0) / 100),
        ("Permit %",                   "Permit_Pct_Source",        markups.get("permit_pct", 2) / 100),
        ("OH&P %",                     "OHP_Pct_Source",           markups.get("ohp_pct", 15) / 100),
    ]
    for i, (label, name, val) in enumerate(markup_panel):
        r = 6 + i
        setval(ws, r, 2, label, font=F_LABEL(), align=Alignment(horizontal="right"))
        c = setval(ws, r, 3, val, fmt=PCT1, font=F_BODY(), align=Alignment(horizontal="right"))
        c.fill = fill("DDF7E0")
        define_name(wb, name, f"Budget_Summary!$C${r}")

    # Cost by CSI division
    section_header(ws, 14, "COST BY CSI DIVISION", span=8)
    cols = ["Div", "Name", "Material (pre-tax)", "Labor", "Equipment", "Subcontract", "Total"]
    for c, h in enumerate(cols, start=1):
        cell = ws.cell(15, c, h)
        cell.fill = fill(NAVY); cell.font = F_H2()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    divs = sorted({str(it.get("division", "")).strip() for it in items if str(it.get("division", "")).strip()},
                  key=lambda x: x.zfill(3))
    rng_div   = "Detail_Div"
    rng_mat   = "Detail_Mat"
    rng_lab   = "Detail_Lab"
    rng_equip = "Detail_Equip"
    rng_sub   = "Detail_Sub"

    first_div_row = 16
    for i, div in enumerate(divs):
        r = first_div_row + i
        setval(ws, r, 1, div, font=F_BODY(), align=Alignment(horizontal="center"))
        setval(ws, r, 2, DIV_NAMES.get(div, ""), font=F_BODY(), align=Alignment(horizontal="left", indent=1))
        ws.cell(r, 3, f'=SUMIF({rng_div},A{r},{rng_mat})')
        ws.cell(r, 4, f'=SUMIF({rng_div},A{r},{rng_lab})')
        ws.cell(r, 5, f'=SUMIF({rng_div},A{r},{rng_equip})')
        ws.cell(r, 6, f'=SUMIF({rng_div},A{r},{rng_sub})')
        ws.cell(r, 7, f"=SUM(C{r}:F{r})")
        for c in range(3, 8):
            ws.cell(r, c).number_format = CURRENCY
            ws.cell(r, c).border = BORDER
        ws.cell(r, 1).border = BORDER; ws.cell(r, 2).border = BORDER
        if i % 2 == 1:
            for c in range(1, 8):
                ws.cell(r, c).fill = fill(ROW_ZEBRA)
    last_div_row = first_div_row + len(divs) - 1

    # Direct cost totals
    direct_row = last_div_row + 1
    setval(ws, direct_row, 1, "Direct Cost", font=F_TOTAL(), align=Alignment(horizontal="right"), fill_rgb=LIGHT_BLUE)
    setval(ws, direct_row, 2, "", fill_rgb=LIGHT_BLUE)
    for col, letter in ((3, "C"), (4, "D"), (5, "E"), (6, "F"), (7, "G")):
        ws.cell(direct_row, col,
                f"=SUM({letter}{first_div_row}:{letter}{last_div_row})").number_format = CURRENCY
        ws.cell(direct_row, col).font = F_TOTAL()
        ws.cell(direct_row, col).fill = fill(LIGHT_BLUE)

    # Markup waterfall (refer back to the markup source panel)
    wf_start = direct_row + 2
    section_header(ws, wf_start, "MARKUP WATERFALL", span=8)
    setval(ws, wf_start + 1, 1, "Step", font=F_H2(), fill_rgb=GRAY, align=Alignment(horizontal="center"))
    setval(ws, wf_start + 1, 5, "Rate", font=F_H2(), fill_rgb=GRAY, align=Alignment(horizontal="center"))
    setval(ws, wf_start + 1, 6, "Amount", font=F_H2(), fill_rgb=GRAY, align=Alignment(horizontal="center"))
    setval(ws, wf_start + 1, 7, "Running", font=F_H2(), fill_rgb=GRAY, align=Alignment(horizontal="center"))

    direct_F = f"G{direct_row}"
    total_material_ref = f"C{direct_row}"

    r = wf_start + 2
    def line(label, formula, rate=None, bold=False, fill_rgb=None):
        nonlocal r
        setval(ws, r, 1, label, font=F_TOTAL() if bold else F_LABEL(),
               align=Alignment(horizontal="right"), fill_rgb=fill_rgb)
        if rate is not None:
            setval(ws, r, 5, rate, fmt=PCT1, font=F_BODY(),
                   align=Alignment(horizontal="right"), fill_rgb=fill_rgb)
        c = setval(ws, r, 6, formula, fmt=CURRENCY,
                   font=F_TOTAL() if bold else F_BODY(),
                   align=Alignment(horizontal="right"), fill_rgb=fill_rgb)
        rr = r
        r += 1
        return rr

    line("Direct Cost",                f"={direct_F}", bold=True)
    subtotal_F = f"F{r-1}"
    tax_r = line("Material Sales Tax (materials only)",
                 f"={total_material_ref}*Material_Tax_Pct_Source",
                 rate=f"=Material_Tax_Pct_Source")
    setval(ws, r, 1, "Subtotal", font=F_TOTAL(),
           align=Alignment(horizontal="right"), fill_rgb=LIGHT_GRAY)
    setval(ws, r, 6, f"={subtotal_F}+F{tax_r}", fmt=CURRENCY,
           font=F_TOTAL(), align=Alignment(horizontal="right"),
           fill_rgb=LIGHT_GRAY)
    subtotal_F = f"F{r}"; r += 1

    for key_name, label in [
        ("GC_Markup_Pct",        "General Conditions"),
        ("Contingency_Pct",      "Contingency / Escalation"),
        ("Insurance_Pct",        "Insurance (GL + Builder's Risk)"),
        ("Bond_Pct",             "Payment & Performance Bond"),
        ("Permit_Pct",           "Permits & Fees"),
        ("OHP_Pct",              "Overhead & Profit"),
    ]:
        line_r = line(label, f"={subtotal_F}*{key_name}", rate=f"={key_name}")
        setval(ws, r, 1, "Subtotal", font=F_TOTAL(),
               align=Alignment(horizontal="right"), fill_rgb=LIGHT_GRAY)
        setval(ws, r, 6, f"={subtotal_F}+F{line_r}", fmt=CURRENCY,
               font=F_TOTAL(), align=Alignment(horizontal="right"),
               fill_rgb=LIGHT_GRAY)
        subtotal_F = f"F{r}"; r += 1

    # BID TOTAL highlight
    bid_r = r
    ws.merge_cells(start_row=bid_r, start_column=1, end_row=bid_r, end_column=5)
    cell = ws.cell(bid_r, 1, "BID TOTAL")
    cell.font = F_BIG_TOTAL(); cell.fill = fill(NAVY)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    c = ws.cell(bid_r, 6, f"={subtotal_F}")
    c.number_format = CURRENCY0; c.font = F_BIG_TOTAL(); c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[bid_r].height = 28

    # Track Bid Total cell for cross-sheet refs (Sources & Uses pulls C8 ← Budget_Summary!F<bid_r>)
    define_name(wb, "Bid_Total", f"Budget_Summary!$F${bid_r}")

    widths = [4, 32, 14, 14, 14, 16, 16, 4]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    return bid_r


def build_sov(wb, items, markups, logo, schedule_months):
    """Schedule of Values (AIA G703-style) — by CSI division + markup lines."""
    ws = wb.create_sheet("Schedule_of_Values")
    add_logo(ws, logo)
    page_header(ws, "Schedule of Values", "AIA G703-style scheduled values", span=12)

    # Header
    headers = ["Item #", "Division", "Description", "Scheduled Value",
               "Work Previous", "Work This Period", "Materials Stored",
               "Total Completed + Stored", "% Complete",
               "Balance to Finish", "Retainage", "Pay This Period"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(6, c, h)
        cell.fill = fill(NAVY); cell.font = F_H2()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[6].height = 36

    # Build items: each CSI division gets one line; then markup lines
    divs = sorted({str(it.get("division", "")).strip() for it in items if str(it.get("division", "")).strip()},
                  key=lambda x: x.zfill(3))

    item_no = 1
    r = 7
    for div in divs:
        ws.cell(r, 1, f"{item_no:03d}")
        ws.cell(r, 2, div)
        ws.cell(r, 3, DIV_NAMES.get(div, ""))
        ws.cell(r, 4, f'=SUMIF(Detail_Div,A{r}*1,Detail_Total)' if div.lstrip("0") else "")
        # use A{r}+0 trick (text-to-number) — actually our Detail Div column is text
        ws.cell(r, 4, f'=SUMIF(Detail_Div,B{r},Detail_Total)')
        ws.cell(r, 5, 0); ws.cell(r, 6, 0); ws.cell(r, 7, 0)         # Editable green
        ws.cell(r, 8, f"=E{r}+F{r}+G{r}")
        ws.cell(r, 9, f"=IF(D{r}=0,0,H{r}/D{r})")
        ws.cell(r, 10, f"=D{r}-H{r}")
        ws.cell(r, 11, f"=H{r}*Retainage_Pct")
        ws.cell(r, 12, f"=F{r}+G{r}-(F{r}+G{r})*Retainage_Pct")
        # styling
        for c in (4, 5, 6, 7, 8, 10, 11, 12):
            ws.cell(r, c).number_format = CURRENCY
            ws.cell(r, c).border = BORDER
        ws.cell(r, 9).number_format = PCT1
        ws.cell(r, 9).border = BORDER
        for c in (1, 2, 3):
            ws.cell(r, c).border = BORDER
        # editable cells green
        for c in (5, 6, 7):
            ws.cell(r, c).fill = fill("DDF7E0")
        ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical="center")
        item_no += 1
        r += 1

    # Markup waterfall lines (Sales Tax + GCs + Cont + Ins + Bond + Permit + OH&P)
    # Each line = its computed amount from Budget Summary as a static reference for the SOV
    markup_lines = [
        ("Material Sales Tax",  "=SUMIF(Detail_Div,B{},Detail_Mat)*Material_Tax_Pct_Source"),
        # NOTE: cannot easily reuse SUMIF; we'll just pull from Budget Summary cells
    ]
    # Simpler: pull all markup amounts via dedicated formulas referencing Budget_Summary panel
    # The markup amounts live at specific cells in Budget_Summary (column F).
    # We'll add them as one-off lines referencing those cells directly.
    summary_markups = [
        ("Material Sales Tax (materials only)",      "BS_Tax_Cell"),
        ("General Conditions",                       "BS_GC_Cell"),
        ("Contingency / Escalation",                 "BS_Cont_Cell"),
        ("Insurance (GL + Builder's Risk)",          "BS_Ins_Cell"),
        ("Payment & Performance Bond",               "BS_Bond_Cell"),
        ("Permits & Fees",                           "BS_Permit_Cell"),
        ("Overhead & Profit",                        "BS_OHP_Cell"),
    ]
    for label, _ in summary_markups:
        # We'll just compute the markup as percentage × direct subtotal locally
        # to keep SOV self-contained.
        pass

    # Skip the per-line markup detail (too complex to cross-link cleanly); use
    # an aggregate "Markups / GR / OH&P" rollup line so SOV total = Bid Total.
    setval(ws, r, 1, f"{item_no:03d}", align=Alignment(horizontal="center"))
    setval(ws, r, 2, "MK")
    setval(ws, r, 3, "Markups / GRs / Sales Tax / Insurance / Permit / OH&P (per Budget Summary)",
           align=Alignment(wrap_text=True, vertical="center"))
    # Scheduled value = Bid Total − sum of div Scheduled values above
    ws.cell(r, 4, f"=Bid_Total-SUM(D7:D{r-1})")
    ws.cell(r, 5, 0); ws.cell(r, 6, 0); ws.cell(r, 7, 0)
    ws.cell(r, 8, f"=E{r}+F{r}+G{r}")
    ws.cell(r, 9, f"=IF(D{r}=0,0,H{r}/D{r})")
    ws.cell(r, 10, f"=D{r}-H{r}")
    ws.cell(r, 11, f"=H{r}*Retainage_Pct")
    ws.cell(r, 12, f"=F{r}+G{r}-(F{r}+G{r})*Retainage_Pct")
    for c in (4, 5, 6, 7, 8, 10, 11, 12):
        ws.cell(r, c).number_format = CURRENCY
        ws.cell(r, c).border = BORDER
    ws.cell(r, 9).number_format = PCT1; ws.cell(r, 9).border = BORDER
    for c in (1, 2, 3):
        ws.cell(r, c).border = BORDER
    for c in (5, 6, 7):
        ws.cell(r, c).fill = fill("DDF7E0")
    last_data = r
    r += 1

    # TOTAL row
    setval(ws, r, 1, "TOTAL", font=F_BIG_TOTAL(), fill_rgb=NAVY,
           align=Alignment(horizontal="center", vertical="center"))
    setval(ws, r, 2, "", fill_rgb=NAVY)
    setval(ws, r, 3, "GRAND TOTAL CONTRACT", font=F_BIG_TOTAL(), fill_rgb=NAVY,
           align=Alignment(horizontal="left", indent=1, vertical="center"))
    for c, col in ((4, "D"), (5, "E"), (6, "F"), (7, "G"), (8, "H"), (10, "J"), (11, "K"), (12, "L")):
        ws.cell(r, c, f"=SUM({col}7:{col}{last_data})").number_format = CURRENCY
        ws.cell(r, c).font = F_BIG_TOTAL(); ws.cell(r, c).fill = fill(NAVY)
    ws.cell(r, 9, f"=IF(D{r}=0,0,H{r}/D{r})").number_format = PCT1
    ws.cell(r, 9).font = F_BIG_TOTAL(); ws.cell(r, 9).fill = fill(NAVY)
    ws.row_dimensions[r].height = 26

    # Column widths
    widths = [8, 8, 36, 16, 14, 14, 14, 16, 11, 16, 12, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A7"


def build_draw_schedule(wb, items, markups, logo, schedule_months):
    """Monthly Draw Schedule using a typical residential phasing matrix."""
    ws = wb.create_sheet("Draw_Schedule")
    add_logo(ws, logo)
    page_header(ws, "Draw Schedule", "Monthly cash-flow forecast (lender view)", span=2 + schedule_months + 2)

    headers = ["Div", "Description", "Scheduled Value"] + [f"M{i+1}" for i in range(schedule_months)] + ["Check Total", "Cum %"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(6, c, h)
        cell.fill = fill(NAVY); cell.font = F_H2()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[6].height = 26

    phasing = phasing_for_n(schedule_months)
    divs = sorted({str(it.get("division", "")).strip() for it in items if str(it.get("division", "")).strip()},
                  key=lambda x: x.zfill(3))

    r = 7
    # Pre-compute scheduled values by division using SUMIF references
    for div in divs:
        ws.cell(r, 1, div)
        ws.cell(r, 2, DIV_NAMES.get(div, ""))
        ws.cell(r, 3, f'=SUMIF(Detail_Div,A{r},Detail_Total)')
        curve = phasing.get(div, [100 / schedule_months] * schedule_months)
        # If the division has its own curve, use it; else even spread
        for m in range(schedule_months):
            ws.cell(r, 4 + m, f"=C{r}*{curve[m]/100:.6f}")
        check_col = 4 + schedule_months
        ws.cell(r, check_col, f"=SUM({get_column_letter(4)}{r}:{get_column_letter(3+schedule_months)}{r})")
        ws.cell(r, check_col + 1, f"=IF(Bid_Total=0,0,SUM({get_column_letter(4)}{r}:{get_column_letter(3+schedule_months)}{r})/Bid_Total)")
        # formatting
        for c in range(3, check_col + 1):
            ws.cell(r, c).number_format = CURRENCY
            ws.cell(r, c).border = BORDER
        ws.cell(r, check_col + 1).number_format = PCT1
        ws.cell(r, check_col + 1).border = BORDER
        for c in (1, 2):
            ws.cell(r, c).border = BORDER
        if (r - 7) % 2 == 1:
            for c in range(1, check_col + 2):
                ws.cell(r, c).fill = fill(ROW_ZEBRA)
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="center")
        r += 1

    last_div_row = r - 1

    # Markups + GRs row (spread across the duration on a different curve)
    # Use even spread for GCs, front-loaded for Permit, back-loaded for Retainage release
    # Simpler: spread the markup-block proportionally to direct draw % per month
    setval(ws, r, 1, "MK")
    setval(ws, r, 2, "Markups / GRs / Sales Tax / Insurance / Permit / OH&P",
           align=Alignment(wrap_text=True, vertical="center"))
    ws.cell(r, 3, f"=Bid_Total-SUM(C7:C{last_div_row})")
    for m in range(schedule_months):
        col = get_column_letter(4 + m)
        # proportional to monthly direct draw
        ws.cell(r, 4 + m,
                f"=IFERROR(C{r} * SUM({col}7:{col}{last_div_row}) / SUM(C7:C{last_div_row}), 0)")
    check_col = 4 + schedule_months
    ws.cell(r, check_col, f"=SUM({get_column_letter(4)}{r}:{get_column_letter(3+schedule_months)}{r})")
    ws.cell(r, check_col + 1, f"=IF(Bid_Total=0,0,SUM({get_column_letter(4)}{r}:{get_column_letter(3+schedule_months)}{r})/Bid_Total)")
    for c in range(3, check_col + 1):
        ws.cell(r, c).number_format = CURRENCY
        ws.cell(r, c).border = BORDER
    ws.cell(r, check_col + 1).number_format = PCT1
    ws.cell(r, check_col + 1).border = BORDER
    for c in (1, 2):
        ws.cell(r, c).border = BORDER
    last_data = r
    r += 1

    # MONTHLY TOTAL row
    setval(ws, r, 1, "TOTAL", font=F_BIG_TOTAL(), fill_rgb=NAVY,
           align=Alignment(horizontal="center", vertical="center"))
    setval(ws, r, 2, "Monthly Total Draw", font=F_BIG_TOTAL(), fill_rgb=NAVY,
           align=Alignment(horizontal="left", indent=1, vertical="center"))
    ws.cell(r, 3, f"=SUM(C7:C{last_data})").number_format = CURRENCY
    ws.cell(r, 3).font = F_BIG_TOTAL(); ws.cell(r, 3).fill = fill(NAVY)
    for m in range(schedule_months):
        col = get_column_letter(4 + m)
        ws.cell(r, 4 + m, f"=SUM({col}7:{col}{last_data})").number_format = CURRENCY
        ws.cell(r, 4 + m).font = F_BIG_TOTAL(); ws.cell(r, 4 + m).fill = fill(NAVY)
    ws.cell(r, check_col, f"=SUM(D{r}:{get_column_letter(3+schedule_months)}{r})").number_format = CURRENCY
    ws.cell(r, check_col).font = F_BIG_TOTAL(); ws.cell(r, check_col).fill = fill(NAVY)
    ws.cell(r, check_col + 1, f"=IF(Bid_Total=0,0,{get_column_letter(check_col)}{r}/Bid_Total)").number_format = PCT1
    ws.cell(r, check_col + 1).font = F_BIG_TOTAL(); ws.cell(r, check_col + 1).fill = fill(NAVY)
    ws.row_dimensions[r].height = 24
    total_row = r
    r += 1

    # CUMULATIVE row
    setval(ws, r, 1, "CUM", font=F_TOTAL(), fill_rgb=LIGHT_BLUE,
           align=Alignment(horizontal="center"))
    setval(ws, r, 2, "Cumulative Draw", font=F_TOTAL(), fill_rgb=LIGHT_BLUE,
           align=Alignment(horizontal="left", indent=1))
    ws.cell(r, 3, f"=C{total_row}").number_format = CURRENCY
    ws.cell(r, 3).fill = fill(LIGHT_BLUE); ws.cell(r, 3).font = F_TOTAL()
    for m in range(schedule_months):
        col = get_column_letter(4 + m)
        prev = get_column_letter(4 + m - 1)
        if m == 0:
            ws.cell(r, 4 + m, f"={col}{total_row}")
        else:
            ws.cell(r, 4 + m, f"={prev}{r}+{col}{total_row}")
        ws.cell(r, 4 + m).number_format = CURRENCY
        ws.cell(r, 4 + m).fill = fill(LIGHT_BLUE); ws.cell(r, 4 + m).font = F_TOTAL()
    cum_row = r
    r += 1

    # CUM % row
    setval(ws, r, 1, "%", font=F_TOTAL(), fill_rgb=LIGHT_BLUE,
           align=Alignment(horizontal="center"))
    setval(ws, r, 2, "Cumulative %", font=F_TOTAL(), fill_rgb=LIGHT_BLUE,
           align=Alignment(horizontal="left", indent=1))
    setval(ws, r, 3, "", fill_rgb=LIGHT_BLUE)
    for m in range(schedule_months):
        col = get_column_letter(4 + m)
        ws.cell(r, 4 + m, f"=IF(Bid_Total=0,0,{col}{cum_row}/Bid_Total)").number_format = PCT1
        ws.cell(r, 4 + m).fill = fill(LIGHT_BLUE); ws.cell(r, 4 + m).font = F_TOTAL()

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 15
    for m in range(schedule_months):
        ws.column_dimensions[get_column_letter(4 + m)].width = 13
    ws.column_dimensions[get_column_letter(4 + schedule_months)].width = 15
    ws.column_dimensions[get_column_letter(5 + schedule_months)].width = 10
    ws.freeze_panes = "D7"

    # Add Month-to-Date labels in row 5 above the M# header
    for m in range(schedule_months):
        setval(ws, 5, 4 + m,
               f"=EDATE(Start_Date, {m})",
               fmt='mmm yyyy', font=F_LABEL(),
               align=Alignment(horizontal="center"))
    setval(ws, 5, 3, "Draw amounts in $; phasing per a typical FL coastal SFR schedule.",
           font=F_NOTE(), align=Alignment(horizontal="left"))


def build_timeline(wb, items, logo, schedule_months):
    ws = wb.create_sheet("Construction_Timeline")
    add_logo(ws, logo)
    page_header(ws, "Construction Timeline", "Gantt-style phasing by major work category", span=2 + schedule_months)

    phases = [
        ("Mobilization, permits, layout, dewatering",                  "31", 1, 2),
        ("Pile driving, grade-beam excavation",                        "31", 2, 2),
        ("Foundation: grade beams, slab, termite, flood vents rough",  "03", 3, 2),
        ("Concrete columns + masonry bearing walls (lower)",           "04", 3, 3),
        ("Tie beams + roof trusses + sheathing",                       "06", 4, 3),
        ("Roof dry-in (NOA membrane) + parapet flashing",              "07", 5, 2),
        ("Window/door install (impact-rated)",                          "08", 6, 2),
        ("MEP rough-in (plumbing, HVAC, electrical, gas)",              "22", 5, 4),
        ("Insulation (R-30 closed-cell foam) + drywall",               "07", 7, 2),
        ("Exterior stucco + paint",                                    "09", 7, 2),
        ("Interior finishes (tile, flooring, paint)",                  "09", 8, 3),
        ("Cabinetry, millwork, appliances",                            "06", 9, 2),
        ("MEP trim + fixtures + elevator install",                     "23", 9, 2),
        ("Pool shell + equipment + barrier",                           "13", 5, 4),
        ("Site finishes (driveway, walks, fence, landscape)",          "32", 10, 2),
        ("Punch list, final clean, Certificate of Occupancy",          "01", 11, 1),
    ]

    # Header row
    setval(ws, 6, 1, "Phase", font=F_H2(), fill_rgb=NAVY,
           align=Alignment(horizontal="center", vertical="center"))
    setval(ws, 6, 2, "Description", font=F_H2(), fill_rgb=NAVY,
           align=Alignment(horizontal="center", vertical="center"))
    for m in range(schedule_months):
        c = setval(ws, 6, 3 + m, f"=EDATE(Start_Date,{m})",
                   fmt='mmm yyyy', font=F_H2(), fill_rgb=NAVY,
                   align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[6].height = 28

    # Phase rows
    for i, (desc, div, start_m, dur) in enumerate(phases):
        r = 7 + i
        setval(ws, r, 1, f"P{i+1:02d}", font=F_LABEL(), align=Alignment(horizontal="center"))
        setval(ws, r, 2, desc, font=F_BODY(),
               align=Alignment(wrap_text=True, vertical="center", indent=1))
        for m in range(schedule_months):
            month_num = m + 1
            if start_m <= month_num < start_m + dur:
                c = ws.cell(r, 3 + m, "█")
                c.fill = fill(NAVY); c.font = Font(color=NAVY)
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c = ws.cell(r, 3 + m, "")
                c.fill = fill(ROW_ZEBRA if i % 2 else WHITE)
            c.border = BORDER
        for c in (1, 2):
            ws.cell(r, c).border = BORDER
            if i % 2:
                ws.cell(r, c).fill = fill(ROW_ZEBRA)

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 52
    for m in range(schedule_months):
        ws.column_dimensions[get_column_letter(3 + m)].width = 11
    ws.freeze_panes = "C7"


def build_scope(wb, items, cfg, logo):
    ws = wb.create_sheet("Scope_of_Work")
    ws.sheet_view.showGridLines = False
    add_logo(ws, logo)
    page_header(ws, "Scope of Work", "Inclusions / exclusions by CSI division", span=8)

    section_header(ws, 5, "PROJECT UNDERSTANDING", span=8)
    para = (f'{cfg["company"]["name"]} ({cfg["company"].get("license","")}) proposes to construct '
            f'the {cfg["project"]["type"]} at {cfg["project"]["address"]}, {cfg["project"]["city_state_zip"]}. '
            f'The project is in FEMA Flood Zone {cfg["project"]["flood_zone"]} (BFE {cfg["project"]["bfe_ft"]} ft) '
            f'and the Windborne Debris Region (V_ult {cfg["project"]["vult_mph"]} mph, Exp {cfg["project"]["exposure"]}). '
            f'Foundation = driven 10″ timber piles + reinforced concrete grade beams; superstructure = '
            f'CMU bearing walls + pre-engineered wood trusses, Type V-B, fully hurricane-strapped. '
            f'Conditioned area ≈ {cfg["project"]["conditioned_sf"]:,} SF; total gross under roof ≈ '
            f'{cfg["project"]["gross_under_roof_sf"]:,} SF over 4 plates (garage / 1st / 2nd / roof deck) with elevator.')
    ws.merge_cells("B6:H8")
    c = ws.cell(6, 2, para)
    c.font = F_BODY()
    c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    ws.row_dimensions[6].height = 36

    # Group items by division and list as inclusions
    divs = sorted({str(it.get("division", "")).strip() for it in items if str(it.get("division", "")).strip()},
                  key=lambda x: x.zfill(3))
    r = 10
    section_header(ws, r, "INCLUSIONS BY CSI DIVISION", span=8)
    r += 1
    for div in divs:
        div_items = [it for it in items if str(it.get("division","")).strip() == div]
        if not div_items:
            continue
        # division header
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        c = ws.cell(r, 2, f"Div {div} — {DIV_NAMES.get(div, '')}")
        c.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        c.fill = fill(LIGHT_BLUE)
        c.alignment = Alignment(horizontal="left", indent=1)
        r += 1
        # items
        for it in div_items:
            qty = _num(it.get("qty"))
            unit = (it.get("unit") or "").strip()
            qstr = f" ({qty:,.0f} {unit})" if qty and unit not in ("LS", "ALLOW", "EA") else (
                f" ({qty:,.0f} {unit})" if qty and unit == "EA" else "")
            line = f"• {it.get('item','')}{qstr} — {it.get('description','')}"
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            c = ws.cell(r, 2, line)
            c.font = F_BODY()
            c.alignment = Alignment(wrap_text=True, vertical="top", indent=2)
            ws.row_dimensions[r].height = max(15, min(45, 15 + len(line) // 90 * 12))
            r += 1
        r += 1   # spacer

    # Exclusions
    section_header(ws, r, "EXCLUSIONS (NOT included in this scope)", span=8, fill_rgb=GRAY)
    r += 1
    exclusions = [
        "Land acquisition costs",
        "Owner-furnished items (refrigerator, washer/dryer, free-standing furniture, art, AV equipment)",
        "Architectural & engineering design fees (covered under soft costs)",
        "Surveys and FEMA Elevation Certificate (post-construction surveyor — owner's responsibility)",
        "Threshold special inspector fees (if applicable; allowance carried for budget transparency)",
        "Impact / school / transportation / utility impact fees (paid by owner outside contract)",
        "Generator and standby power (not shown on drawings — add via Alternate)",
        "Existing structure demolition / hazmat / abatement (vacant-lot basis assumed)",
        "Hurricane-shutter alternative (impact-rated openings are base; shutters available as deduct alternate)",
        "Pool deck heaters, automation, lighting beyond pool sub's standard package",
        "Smart-home / structured AV / security beyond the carried allowances",
        "Site dewatering beyond the carried allowance",
        "Off-site or jurisdictional improvements (sidewalks, curbing, easement work outside lot)",
        "Builder's risk insurance after substantial completion",
        "Maintenance, warranty repairs, or change-of-scope work after CO",
    ]
    for x in exclusions:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        c = ws.cell(r, 2, f"• {x}")
        c.font = F_BODY()
        c.alignment = Alignment(wrap_text=True, indent=2)
        r += 1

    # Clarifications
    r += 1
    section_header(ws, r, "CLARIFICATIONS & QUALIFICATIONS", span=8, fill_rgb=GRAY)
    r += 1
    clarifications = [
        f'Basis of bid: signed/sealed drawings as listed in the Plans index + Structural Calcs (Calc Engineering C-26225); no addenda.',
        f'Pricing valid for 30 days from issuance; material escalation reserved beyond.',
        f'Florida sales tax (6% state + 1% Pinellas surtax = 7%) applied to materials only per FBC Ch 212.',
        f'Florida lien-law (Ch. 713) notice rights reserved.',
        f'Contractor License: {cfg["company"].get("license","[CGC# pending]")}.',
        f'Bond not carried (private SFR — no FL 255.05 trigger). Add via Alternate if lender requires.',
        f'GL + Builder’s Risk insurance carried by GC per Inputs.',
        f'Stairs (3 runs) and frameless glass guardrails per A5.0 "by others" — carried as allowances.',
        f'Pool shell + equipment + barrier carried as allowances under Div 13; final scope per pool sub package.',
        f'Impact-rated openings carried as code-required (FBC WBDR); FL# / NOA approvals to be confirmed at submittal.',
    ]
    for x in clarifications:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        c = ws.cell(r, 2, f"• {x}")
        c.font = F_BODY()
        c.alignment = Alignment(wrap_text=True, indent=2)
        ws.row_dimensions[r].height = max(15, min(45, 15 + len(x) // 90 * 12))
        r += 1

    widths = [2, 60, 14, 14, 14, 14, 14, 14, 4]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_allowances(wb, items, logo):
    ws = wb.create_sheet("Allowances")
    ws.sheet_view.showGridLines = False
    add_logo(ws, logo)
    page_header(ws, "Allowances", "Carry-as-allowance items in the bid", span=6)

    headers = ["Div", "Item", "Description", "Amount", "Basis / Notes"]
    for c, h in enumerate(headers, start=2):
        cell = ws.cell(6, c, h)
        cell.fill = fill(NAVY); cell.font = F_H2()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[6].height = 24

    r = 7
    total_rows = []
    for it in items:
        unit = (it.get("unit") or "").strip().upper()
        if unit in ("ALLOW", "LS"):
            ws.cell(r, 2, str(it.get("division","")).strip())
            ws.cell(r, 3, it.get("item",""))
            ws.cell(r, 4, it.get("description",""))
            qty = _num(it.get("qty"))
            sub = _num(it.get("unit_sub"))
            mat = _num(it.get("unit_mat"))
            lab = _num(it.get("unit_lab"))
            eqp = _num(it.get("unit_equip"))
            amt = qty * (sub + mat + lab + eqp)
            ws.cell(r, 5, amt).number_format = CURRENCY
            ws.cell(r, 6, it.get("notes",""))
            for c in range(2, 7):
                ws.cell(r, c).border = BORDER
            ws.cell(r, 4).alignment = Alignment(wrap_text=True, vertical="center")
            ws.cell(r, 6).alignment = Alignment(wrap_text=True, vertical="center")
            if (r-7) % 2:
                for c in range(2, 7):
                    ws.cell(r, c).fill = fill(ROW_ZEBRA)
            total_rows.append(r)
            r += 1

    # Total
    setval(ws, r, 2, "TOTAL ALLOWANCES", font=F_BIG_TOTAL(), fill_rgb=NAVY,
           align=Alignment(horizontal="right", indent=1))
    setval(ws, r, 3, "", fill_rgb=NAVY)
    setval(ws, r, 4, "", fill_rgb=NAVY)
    if total_rows:
        ws.cell(r, 5, f"=SUM(E{total_rows[0]}:E{total_rows[-1]})").number_format = CURRENCY
    else:
        ws.cell(r, 5, 0).number_format = CURRENCY
    ws.cell(r, 5).font = F_BIG_TOTAL(); ws.cell(r, 5).fill = fill(NAVY)
    setval(ws, r, 6, "", fill_rgb=NAVY)
    ws.row_dimensions[r].height = 24

    widths = [2, 6, 28, 50, 16, 36, 4]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A7"


def build_alternates(wb, logo):
    ws = wb.create_sheet("Alternates_and_Unit_Prices")
    ws.sheet_view.showGridLines = False
    add_logo(ws, logo)
    page_header(ws, "Alternates & Unit Prices", "Owner / lender selection items", span=6)

    section_header(ws, 5, "ALTERNATES", span=6)
    headers = ["#", "Description", "Add / Deduct", "Amount", "Notes"]
    for c, h in enumerate(headers, start=2):
        cell = ws.cell(6, c, h)
        cell.fill = fill(NAVY); cell.font = F_H2()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    alts = [
        ("1", "Pool screen enclosure (in lieu of standard pool barrier)",            "ADD",     30000, "Removed; carries pool enclosure allowance"),
        ("2", "Standby generator + transfer switch (whole-house)",                   "ADD",     22000, "Typical FL coastal upgrade"),
        ("3", "Hurricane-shutter system in lieu of impact-rated openings (entire envelope)", "DEDUCT", -28000, "Code permits either path"),
        ("4", "TPO membrane in lieu of SBS modified bitumen roofing",                "DEDUCT",  -8000, "VE — both NOA approved"),
        ("5", "Site-finished hardwood in lieu of LVT in living/bedroom areas",       "ADD",     18000, "Owner preference"),
        ("6", "Glass shower enclosures upgrade (frameless, all baths)",              "ADD",      9000, "From standard semi-frameless"),
        ("7", "Solar-ready conduit + roof attachments (PV-ready, no panels)",        "ADD",      4500, "Future PV"),
        ("8", "EV charger + Level 2 panel upgrade (per stall above 2)",              "ADD/EA",   2400, "Per added Level 2 location"),
    ]
    for i, (no, desc, ad, amt, notes) in enumerate(alts):
        r = 7 + i
        setval(ws, r, 2, no, align=Alignment(horizontal="center"))
        setval(ws, r, 3, desc, align=Alignment(wrap_text=True, indent=1, vertical="center"))
        setval(ws, r, 4, ad, align=Alignment(horizontal="center"))
        c = setval(ws, r, 5, amt, fmt=CURRENCY,
                   font=F_BODY(),
                   align=Alignment(horizontal="right"))
        if amt < 0:
            c.font = Font(name="Calibri", size=10, color=ACCENT_GREEN, bold=True)
        setval(ws, r, 6, notes, align=Alignment(wrap_text=True, indent=1, vertical="center"))
        if i % 2:
            for c in range(2, 7):
                ws.cell(r, c).fill = fill(ROW_ZEBRA)

    # Unit Prices
    section_header(ws, 18, "UNIT PRICES (Owner add/deduct authorizations)", span=6)
    headers = ["#", "Item", "Unit", "$ / Unit", "Notes"]
    for c, h in enumerate(headers, start=2):
        cell = ws.cell(19, c, h)
        cell.fill = fill(NAVY); cell.font = F_H2()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    units = [
        ("1", "Slab thickness over base 6\" (incremental fill + concrete)",  "per inch",  4000,  "RFI: General Note 4 says 8\""),
        ("2", "Impact window/door (added beyond base count)",                "per SF",    105,   "Avg $100-120/SF installed"),
        ("3", "Timber pile (over-pile beyond 60)",                           "per EA",    1620,  "Mat + driving + cutoff"),
        ("4", "Timber pile length (deeper driving)",                         "per LF",    35,    "Same unit price as base"),
        ("5", "Concrete grade beam concrete (over)",                         "per CY",    280,   "Includes pump + finish"),
        ("6", "Grade-beam rebar (over)",                                     "per ton",   3400,  "Material + labor"),
        ("7", "Dewatering (added)",                                          "per week",  3500,  "Pump + filtration"),
        ("8", "CMU 8\" wall (over base SF)",                                 "per SF",    8,     "Block + grout + reinf + labor"),
    ]
    for i, (no, item, unit, price, notes) in enumerate(units):
        r = 20 + i
        setval(ws, r, 2, no, align=Alignment(horizontal="center"))
        setval(ws, r, 3, item, align=Alignment(wrap_text=True, indent=1, vertical="center"))
        setval(ws, r, 4, unit, align=Alignment(horizontal="center"))
        setval(ws, r, 5, price, fmt=CURRENCY, align=Alignment(horizontal="right"))
        setval(ws, r, 6, notes, align=Alignment(wrap_text=True, indent=1, vertical="center"))
        if i % 2:
            for c in range(2, 7):
                ws.cell(r, c).fill = fill(ROW_ZEBRA)

    widths = [2, 5, 52, 14, 16, 36, 4]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_documents(wb, logo):
    ws = wb.create_sheet("Documents_Checklist")
    ws.sheet_view.showGridLines = False
    add_logo(ws, logo)
    page_header(ws, "Documents Checklist", "Required documents for the loan submission", span=6)

    headers = ["Category", "Document", "Status", "Notes / File"]
    for c, h in enumerate(headers, start=2):
        cell = ws.cell(6, c, h)
        cell.fill = fill(NAVY); cell.font = F_H2()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    docs = [
        ("GC Documents",  "Certified General Contractor license (current)",     "Pending", "Attach copy"),
        ("GC Documents",  "GL Insurance Certificate (with lender as additional insured)", "Pending", "Attach"),
        ("GC Documents",  "Workers' Compensation Insurance Certificate",        "Pending", "Attach"),
        ("GC Documents",  "Builder's Risk Insurance Certificate (during construction)", "Pending", "Attach"),
        ("GC Documents",  "Performance & Payment Bond (if required by lender)", "N/A — private SFR", "Add if lender requires"),
        ("GC Documents",  "W-9 + EIN",                                          "Pending", "Attach"),
        ("Project Docs",  "Signed contract (AIA A101/A102 or custom)",          "Pending", "Borrower + GC"),
        ("Project Docs",  "Notice of Commencement (FL 713.13)",                 "After NTP", "Recorded copy"),
        ("Drawings",      "Architectural set (A0.0–A5.0)",                      "✓ On hand", "Signed/sealed"),
        ("Drawings",      "Structural set (S1.0–S8.0)",                         "✓ Partial", "S6/S7/S8 pending visual takeoff confirm"),
        ("Drawings",      "MEP set (M, E, P, FG)",                              "✓ On hand", "Signed/sealed"),
        ("Drawings",      "Civil / Site plan (ST0.0)",                          "✓ On hand", "Within A-set"),
        ("Engineering",   "Structural calculations (Calc Engineering C-26225)", "✓ On hand", "Sealed 2026-04-22"),
        ("Engineering",   "Geotech report (Central FL Testing #249419)",        "✓ On hand", "Per project basis"),
        ("Engineering",   "Energy calcs (FBC-Energy Form R405)",                "✓ Per project", "Confirm"),
        ("Permits",       "Building permit application",                        "Pending", "AHJ: Treasure Island / Pinellas County"),
        ("Permits",       "Plumbing / Mechanical / Electrical sub-permits",     "Pending", "Per discipline"),
        ("Permits",       "Pool permit",                                        "Pending", "Pool sub package"),
        ("Permits",       "FEMA Elevation Certificate (proposed)",              "Pending", "Pre-construction by owner surveyor"),
        ("Compliance",    "Florida Product Approval / NOA — impact openings",   "RFI",     "FL # to be supplied with submittal"),
        ("Compliance",    "Florida Product Approval / NOA — roofing",           "RFI",     "Per FBC Ch 15"),
        ("Compliance",    "Threshold Inspector engagement letter (if applicable)", "TBD", "3-story / 47 ft borderline FL 553.79"),
        ("Compliance",    "Subterranean termite treatment certificate",         "After application", "FL Memo 173"),
        ("Lender",        "Appraisal (subject-to-completion)",                  "Pending", "Lender-ordered"),
        ("Lender",        "Title commitment",                                   "Pending", "Title company"),
        ("Lender",        "Construction loan agreement",                        "Pending", "Lender"),
        ("Lender",        "Draw request form (AIA G702/G703 or lender)",        "Pending", "First draw at NTP"),
    ]
    for i, (cat, doc, status, notes) in enumerate(docs):
        r = 7 + i
        setval(ws, r, 2, cat, font=F_LABEL(), align=Alignment(horizontal="left", indent=1, vertical="center"))
        setval(ws, r, 3, doc, font=F_BODY(), align=Alignment(wrap_text=True, vertical="center"))
        c = setval(ws, r, 4, status, font=F_BODY(), align=Alignment(horizontal="center"))
        if "✓" in status:
            c.font = Font(name="Calibri", size=10, color=ACCENT_GREEN, bold=True)
        elif "Pending" in status or "RFI" in status or "TBD" in status:
            c.font = Font(name="Calibri", size=10, color=ACCENT_RED, bold=True)
        setval(ws, r, 5, notes, font=F_BODY(), align=Alignment(wrap_text=True, indent=1, vertical="center"))
        if i % 2:
            for c in range(2, 6):
                ws.cell(r, c).fill = fill(ROW_ZEBRA)

    widths = [2, 18, 56, 22, 36, 4]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A7"


# ---------------------------------------------------------------------------
def main():
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    proj = Path(sys.argv[1])
    if not proj.is_dir():
        sys.exit(f"Not a directory: {proj}")

    items   = load_lineitems(proj / "lineitems.csv")
    markups = load_markups(proj / "markups.csv")
    cfg     = load_config(proj / "loan-package-config.json")
    # Logo is optional & configurable: config company.logo, else logo.png in the
    # project folder, else a legacy name. If none found, the workbook still builds.
    logo = None
    cfg_logo = (cfg.get("company", {}) or {}).get("logo")
    candidates = []
    if cfg_logo:
        lp = Path(cfg_logo)
        candidates.append(lp if lp.is_absolute() else (proj / cfg_logo))
    candidates += [proj / "logo.png", proj / "Ideal_Construction_logo.png"]
    for cand in candidates:
        if cand and Path(cand).exists():
            logo = cand
            break
    if logo is None:
        print("Note: no logo found (config company.logo / logo.png) — building without a logo image.")

    wb = Workbook()
    # Remove default Sheet
    wb.remove(wb.active)

    # Build tabs in order
    build_cover(wb, cfg, logo)
    build_inputs(wb, cfg, logo)
    # Budget Detail must build before Summary (Summary defines Bid_Total name)
    detail_start, detail_last = build_budget_detail(wb, items, logo)
    bid_r = build_budget_summary(wb, items, markups, logo, detail_start, detail_last)
    build_sources_uses(wb, cfg, logo)
    build_executive_summary(wb, cfg, logo)
    months = int(cfg["schedule"]["duration_months"])
    build_sov(wb, items, markups, logo, months)
    build_draw_schedule(wb, items, markups, logo, months)
    build_timeline(wb, items, logo, months)
    build_scope(wb, items, cfg, logo)
    build_allowances(wb, items, logo)
    build_alternates(wb, logo)
    build_documents(wb, logo)

    # Reorder tabs for bank-friendly navigation
    order = ["Cover", "Inputs", "Executive Summary", "Sources_and_Uses",
             "Budget_Summary", "Budget_Detail", "Schedule_of_Values",
             "Draw_Schedule", "Construction_Timeline", "Scope_of_Work",
             "Allowances", "Alternates_and_Unit_Prices", "Documents_Checklist"]
    new_order = [wb[name] for name in order if name in wb.sheetnames]
    wb._sheets = new_order

    # Color tabs: navy = bank summary view; gray = supporting detail; green = inputs;
    # light-blue = scope/schedule reference
    TAB_COLORS = {
        "Cover":                       "1F3A6C",   # navy
        "Inputs":                      "2E7D32",   # green (editable)
        "Executive Summary":           "1F3A6C",
        "Sources_and_Uses":            "1F3A6C",
        "Budget_Summary":              "1F3A6C",
        "Budget_Detail":               "595D62",   # gray (detail)
        "Schedule_of_Values":          "1F3A6C",
        "Draw_Schedule":               "1F3A6C",
        "Construction_Timeline":       "5E81AC",   # blue
        "Scope_of_Work":               "5E81AC",
        "Allowances":                  "595D62",
        "Alternates_and_Unit_Prices":  "595D62",
        "Documents_Checklist":         "595D62",
    }
    for n, color in TAB_COLORS.items():
        if n in wb.sheetnames:
            wb[n].sheet_properties.tabColor = color

    # Print setup (landscape, fit-to-width, header+footer) on every sheet
    for n in wb.sheetnames:
        ws = wb[n]
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
        ws.page_margins.left = 0.4; ws.page_margins.right = 0.4
        ws.page_margins.top = 0.5; ws.page_margins.bottom = 0.5
        ws.page_margins.header = 0.2; ws.page_margins.footer = 0.2
        ws.oddHeader.right.text = "&\"Calibri,Bold\"&12&K1F3A6CIDEAL CONSTRUCTION"
        ws.oddHeader.left.text = f"&\"Calibri,Italic\"&10&K595D62{cfg['project']['name']}"
        ws.oddFooter.center.text = "&P / &N"
        ws.oddFooter.right.text = "&\"Calibri,Italic\"&9&K595D62Confidential — Construction Loan Package"

    wb.active = 0  # set Cover as the active sheet

    out = proj / "construction-loan-package.xlsx"
    wb.save(out)
    print(f"Wrote {out}  ({len(items)} line items, {len(wb.sheetnames)} tabs)")


if __name__ == "__main__":
    main()
